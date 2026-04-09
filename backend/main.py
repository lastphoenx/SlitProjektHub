# backend/main.py - FastAPI Backend für SlitProjektHub
from __future__ import annotations
import json
import logging
import sys
import os
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, HTTPException, Request, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

# ── Logging ────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s [%(name)s] %(message)s",
)

# ── Paths ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))
os.chdir(ROOT)  # DB-Pfad Referenz

# ── Src imports ────────────────────────────────────────────────────────────
from src.m07_roles import list_roles_df, load_role, upsert_role, soft_delete_role, function_suggestions
from src.m07_tasks import list_tasks_df, load_task, upsert_task, soft_delete_task
from src.m07_projects import list_projects_df, load_project, upsert_project, soft_delete_project
from src.m07_contexts import list_contexts_df, load_context, upsert_context, soft_delete_context
from src.m03_db import get_session, Project, Role, Task, Context
from src.m08_llm import (providers_available, generate_role_text, generate_summary,
                          try_models_with_messages, get_available_models, AVAILABLE_MODELS,
                          generate_role_details, generate_project_details, generate_context_details,
                          have_key, test_connection, DEFAULT_MODELS)
from src.m10_chat import (save_message, load_history, find_latest_session_for_project, build_project_map,
                           update_message_metadata, delete_message, delete_history, purge_history,
                           save_rag_feedback)
from src.m09_rag import retrieve_relevant_chunks_hybrid, build_rag_context_from_search, deduplicate_results, rag_low_confidence_warning
from src.m01_config import load_user_settings, save_user_settings
from sqlmodel import select
import uuid as _uuid

# ── FastAPI app ────────────────────────────────────────────────────────────
app = FastAPI(
    title="SlitProjektHub API",
    description="Backend API für das Projektmanagement-Tool",
    version="2.0.0",
)

# ── Static files & Templates ───────────────────────────────────────────────
app.mount("/static", StaticFiles(directory=BACKEND_DIR / "static"), name="static")
templates = Jinja2Templates(directory=BACKEND_DIR / "templates")

# ── CORS ───────────────────────────────────────────────────────────────────
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:8501", "http://localhost:8502", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Project types (shared constant) ───────────────────────────────────────
PROJECT_TYPES = ["Intern", "Extern", "Studie", "Prototyp", "Betrieb", "Sonstiges"]

# ── Helper: build project list for templates ───────────────────────────────
def _projects_list() -> list[dict]:
    df = list_projects_df(include_deleted=False)
    if df is None or df.empty:
        return []
    return df.rename(columns={
        "Key": "key", "Titel": "title", "KurzTitel": "short_title",
        "Kürzel": "short_code", "Typ": "type", "Beschreibung": "description",
    }).to_dict("records")

# ── Helper: stats for dashboard ────────────────────────────────────────────
def _dashboard_stats() -> dict:
    stats = {"projects": 0, "roles": 0, "tasks": 0, "contexts": 0}
    try:
        with get_session() as ses:
            stats["projects"] = ses.exec(select(Project).where(Project.is_deleted == False)).all().__len__()
            stats["roles"] = ses.exec(select(Role).where(Role.is_deleted == False)).all().__len__()
            stats["tasks"] = ses.exec(select(Task).where(Task.is_deleted == False)).all().__len__()
            stats["contexts"] = ses.exec(select(Context).where(Context.is_deleted == False)).all().__len__()
    except Exception:
        pass
    return stats


# ════════════════════════════════════════════════════════════════════════════
# HTML VIEWS (Jinja2 + HTMX)
# ════════════════════════════════════════════════════════════════════════════

@app.get("/", response_class=HTMLResponse)
async def root_redirect():
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/dashboard")


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "active_page": "dashboard",
        "stats": _dashboard_stats(),
    })


# ── Projects HTML views ────────────────────────────────────────────────────

@app.get("/projects", response_class=HTMLResponse)
async def projects_page(request: Request):
    return templates.TemplateResponse("projects/index.html", {
        "request": request,
        "active_page": "projects",
        "projects": _projects_list(),
    })


@app.get("/projects/new", response_class=HTMLResponse)
async def projects_new_form(request: Request):
    """HTMX partial: Neues-Projekt-Formular"""
    return templates.TemplateResponse("projects/_form.html", {
        "request": request,
        "project": None,
        "body_text": "",
        "project_types": PROJECT_TYPES,
    })


@app.get("/projects/{key}/edit", response_class=HTMLResponse)
async def projects_edit_form(request: Request, key: str):
    """HTMX partial: Bearbeiten-Formular für bestehendes Projekt"""
    proj, body = load_project(key)
    if not proj:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    return templates.TemplateResponse("projects/_form.html", {
        "request": request,
        "project": proj,
        "body_text": body,
        "project_types": PROJECT_TYPES,
    })


@app.get("/projects/{key}/confirm-delete", response_class=HTMLResponse)
async def projects_confirm_delete(request: Request, key: str):
    """HTMX partial: Lösch-Bestätigung"""
    proj, _ = load_project(key)
    if not proj:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    return templates.TemplateResponse("projects/_confirm_delete.html", {
        "request": request,
        "project": proj,
    })


@app.post("/projects", response_class=HTMLResponse)
async def projects_create(
    request: Request,
    title: str = Form(...),
    description: str = Form(...),
    short_title: str = Form(""),
    short_code: str = Form(""),
    type: str = Form(""),
    body_text: str = Form(""),
):
    """HTMX: Projekt erstellen → gibt aktualisierten table-body zurück"""
    try:
        upsert_project(
            title=title.strip(),
            type_name=type.strip() or None,
            body_text=body_text,
            short_title=short_title.strip() or None,
            short_code=short_code.strip() or None,
            description=description.strip(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    response = templates.TemplateResponse("projects/_table.html", {
        "request": request,
        "projects": _projects_list(),
    })
    response.headers["HX-Toast"] = json.dumps({"message": f'Projekt "{title}" erstellt', "type": "success"})
    return response


@app.put("/projects/{key}", response_class=HTMLResponse)
async def projects_update(
    request: Request,
    key: str,
    title: str = Form(...),
    description: str = Form(...),
    short_title: str = Form(""),
    short_code: str = Form(""),
    type: str = Form(""),
    body_text: str = Form(""),
):
    """HTMX: Projekt aktualisieren → gibt aktualisierten table-body zurück"""
    proj, _ = load_project(key)
    if not proj:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    try:
        upsert_project(
            key=key,
            title=title.strip(),
            type_name=type.strip() or None,
            body_text=body_text,
            short_title=short_title.strip() or None,
            short_code=short_code.strip() or None,
            description=description.strip(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    response = templates.TemplateResponse("projects/_table.html", {
        "request": request,
        "projects": _projects_list(),
    })
    response.headers["HX-Toast"] = json.dumps({"message": f'Projekt "{title}" gespeichert', "type": "success"})
    return response


@app.delete("/projects/{key}", response_class=HTMLResponse)
async def projects_delete(key: str):
    """HTMX: Projekt soft-delete → gibt leeren String zurück (Row verschwindet)"""
    success = soft_delete_project(key)
    if not success:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    # Return empty string — HTMX swaps out the row with nothing
    response = HTMLResponse(content="")
    response.headers["HX-Toast"] = json.dumps({"message": "Projekt gelöscht", "type": "success"})
    return response


# ── Roles helper ───────────────────────────────────────────────────────────

def _roles_list() -> list[dict]:
    df = list_roles_df(include_deleted=False)
    if df is None or df.empty:
        return []
    return df.rename(columns={
        "Key": "key",
        "Rollenbezeichnung": "title",
        "Rollenkürzel": "short_code",
        "Beschreibung": "description",
        "Hauptverantwortlichkeiten": "responsibilities",
        "Qualifikationen": "qualifications",
        "Expertise": "expertise",
    }).to_dict("records")


def _role_full(key: str) -> dict | None:
    """Lädt eine Rolle mit allen Feldern (ungekürzt) für das Bearbeitungs-Formular."""
    role_obj, body = load_role(key)
    if not role_obj:
        return None
    return {
        "key": role_obj.key,
        "title": role_obj.title,
        "short_code": role_obj.short_code or "",
        "description": role_obj.description or "",
        "responsibilities": role_obj.responsibilities or "",
        "qualifications": role_obj.qualifications or "",
        "expertise": role_obj.expertise or "",
        "rag_status": role_obj.rag_status,
        "body_text": body,
    }


# ── Roles HTML views ────────────────────────────────────────────────────────

@app.get("/roles", response_class=HTMLResponse)
async def roles_page(request: Request):
    return templates.TemplateResponse("roles/index.html", {
        "request": request,
        "active_page": "roles",
        "roles": _roles_list(),
    })


@app.get("/roles/new", response_class=HTMLResponse)
async def roles_new_form(request: Request):
    """HTMX partial: Neue-Rolle-Formular"""
    return templates.TemplateResponse("roles/_form.html", {
        "request": request,
        "role": None,
        "body_text": "",
        "suggestions": function_suggestions(),
    })


@app.get("/roles/{key}/edit", response_class=HTMLResponse)
async def roles_edit_form(request: Request, key: str):
    """HTMX partial: Bearbeiten-Formular"""
    role = _role_full(key)
    if not role:
        raise HTTPException(status_code=404, detail="Rolle nicht gefunden")
    return templates.TemplateResponse("roles/_form.html", {
        "request": request,
        "role": role,
        "body_text": role["body_text"],
        "suggestions": function_suggestions(),
    })


@app.get("/roles/{key}/confirm-delete", response_class=HTMLResponse)
async def roles_confirm_delete(request: Request, key: str):
    """HTMX partial: Lösch-Bestätigung"""
    role_obj, _ = load_role(key)
    if not role_obj:
        raise HTTPException(status_code=404, detail="Rolle nicht gefunden")
    return templates.TemplateResponse("roles/_confirm_delete.html", {
        "request": request,
        "role": role_obj,
    })


@app.post("/roles", response_class=HTMLResponse)
async def roles_create(
    request: Request,
    title: str = Form(...),
    short_code: str = Form(""),
    description: str = Form(""),
    responsibilities: str = Form(""),
    qualifications: str = Form(""),
    expertise: str = Form(""),
    body_text: str = Form(""),
):
    """HTMX: Rolle erstellen → gibt aktualisierten table-body zurück"""
    try:
        upsert_role(
            title=title.strip(),
            body_text=body_text,
            short_code=short_code.strip() or None,
            description=description.strip() or None,
            responsibilities=responsibilities.strip() or None,
            qualifications=qualifications.strip() or None,
            expertise=expertise.strip() or None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    response = templates.TemplateResponse("roles/_table.html", {
        "request": request,
        "roles": _roles_list(),
    })
    response.headers["HX-Toast"] = json.dumps({"message": f'Rolle "{title}" erstellt', "type": "success"})
    return response


@app.put("/roles/{key}", response_class=HTMLResponse)
async def roles_update(
    request: Request,
    key: str,
    title: str = Form(...),
    short_code: str = Form(""),
    description: str = Form(""),
    responsibilities: str = Form(""),
    qualifications: str = Form(""),
    expertise: str = Form(""),
    body_text: str = Form(""),
):
    """HTMX: Rolle aktualisieren → gibt aktualisierten table-body zurück"""
    role_obj, _ = load_role(key)
    if not role_obj:
        raise HTTPException(status_code=404, detail="Rolle nicht gefunden")
    try:
        upsert_role(
            key=key,
            title=title.strip(),
            body_text=body_text,
            short_code=short_code.strip() or None,
            description=description.strip() or None,
            responsibilities=responsibilities.strip() or None,
            qualifications=qualifications.strip() or None,
            expertise=expertise.strip() or None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    response = templates.TemplateResponse("roles/_table.html", {
        "request": request,
        "roles": _roles_list(),
    })
    response.headers["HX-Toast"] = json.dumps({"message": f'Rolle "{title}" gespeichert', "type": "success"})
    return response


@app.delete("/roles/{key}", response_class=HTMLResponse)
async def roles_delete(key: str):
    """HTMX: Rolle soft-delete → Row verschwindet"""
    if not soft_delete_role(key):
        raise HTTPException(status_code=404, detail="Rolle nicht gefunden")
    response = HTMLResponse(content="")
    response.headers["HX-Toast"] = json.dumps({"message": "Rolle gelöscht", "type": "success"})
    return response


# ── Tasks helper ─────────────────────────────────────────────────────────

def _tasks_list(role_filter: str | None = None) -> list[dict]:
    df = list_tasks_df(include_deleted=False, include_metadata=True)
    if df is None or df.empty:
        return []
    records = df.rename(columns={
        "Key": "key",
        "Titel": "title",
        "KurzTitel": "short_title",
        "Kürzel": "short_code",
        "Funktion": "function",
        "Beschreibung": "description",
        "Quell-Rolle": "source_role_key",
        "Verantwortlichkeit": "source_responsibility",
    }).to_dict("records")
    if role_filter:
        records = [r for r in records if r.get("source_role_key") == role_filter]
    return records


def _roles_for_select() -> list[dict]:
    """Alle nicht-gelöschten Rollen als {key, title} für Select-Dropdowns."""
    df = list_roles_df(include_deleted=False)
    if df is None or df.empty:
        return []
    return (
        df[["Key", "Rollenbezeichnung"]]
        .rename(columns={"Key": "key", "Rollenbezeichnung": "title"})
        .to_dict("records")
    )


# ── Tasks HTML views ──────────────────────────────────────────────────────

@app.get("/tasks", response_class=HTMLResponse)
async def tasks_page(request: Request, role: str | None = None):
    return templates.TemplateResponse("tasks/index.html", {
        "request": request,
        "active_page": "tasks",
        "tasks": _tasks_list(role_filter=role),
        "roles": _roles_for_select(),
        "active_role_filter": role or "",
    })


@app.get("/tasks/new", response_class=HTMLResponse)
async def tasks_new_form(request: Request, role_key: str = ""):
    return templates.TemplateResponse("tasks/_form.html", {
        "request": request,
        "task": None,
        "body_text": "",
        "roles": _roles_for_select(),
        "preselect_role": role_key,
    })


@app.get("/tasks/{key}/edit", response_class=HTMLResponse)
async def tasks_edit_form(request: Request, key: str):
    task_obj, body = load_task(key)
    if not task_obj:
        raise HTTPException(status_code=404, detail="Aufgabe nicht gefunden")
    task = {
        "key": task_obj.key,
        "title": task_obj.title,
        "short_title": task_obj.short_title or "",
        "short_code": task_obj.short_code or "",
        "description": task_obj.description or "",
        "source_role_key": task_obj.source_role_key or "",
        "source_responsibility": task_obj.source_responsibility or "",
    }
    return templates.TemplateResponse("tasks/_form.html", {
        "request": request,
        "task": task,
        "body_text": body,
        "roles": _roles_for_select(),
        "preselect_role": task_obj.source_role_key or "",
    })


@app.get("/tasks/{key}/confirm-delete", response_class=HTMLResponse)
async def tasks_confirm_delete(request: Request, key: str):
    task_obj, _ = load_task(key)
    if not task_obj:
        raise HTTPException(status_code=404, detail="Aufgabe nicht gefunden")
    return templates.TemplateResponse("tasks/_confirm_delete.html", {
        "request": request,
        "task": task_obj,
    })


@app.post("/tasks", response_class=HTMLResponse)
async def tasks_create(
    request: Request,
    title: str = Form(...),
    short_title: str = Form(""),
    short_code: str = Form(""),
    description: str = Form(""),
    source_role_key: str = Form(""),
    body_text: str = Form(""),
):
    try:
        upsert_task(
            title=title.strip(),
            body_text=body_text,
            short_title=short_title.strip() or None,
            short_code=short_code.strip() or None,
            description=description.strip() or None,
            source_role_key=source_role_key.strip() or None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    response = templates.TemplateResponse("tasks/_table.html", {
        "request": request,
        "tasks": _tasks_list(),
        "roles": _roles_for_select(),
    })
    response.headers["HX-Toast"] = json.dumps({"message": f'Aufgabe "{title}" erstellt', "type": "success"})
    return response


@app.put("/tasks/{key}", response_class=HTMLResponse)
async def tasks_update(
    request: Request,
    key: str,
    title: str = Form(...),
    short_title: str = Form(""),
    short_code: str = Form(""),
    description: str = Form(""),
    source_role_key: str = Form(""),
    body_text: str = Form(""),
):
    task_obj, _ = load_task(key)
    if not task_obj:
        raise HTTPException(status_code=404, detail="Aufgabe nicht gefunden")
    try:
        upsert_task(
            key=key,
            title=title.strip(),
            body_text=body_text,
            short_title=short_title.strip() or None,
            short_code=short_code.strip() or None,
            description=description.strip() or None,
            source_role_key=source_role_key.strip() or None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    response = templates.TemplateResponse("tasks/_table.html", {
        "request": request,
        "tasks": _tasks_list(),
        "roles": _roles_for_select(),
    })
    response.headers["HX-Toast"] = json.dumps({"message": f'Aufgabe "{title}" gespeichert', "type": "success"})
    return response


@app.delete("/tasks/{key}", response_class=HTMLResponse)
async def tasks_delete(key: str):
    if not soft_delete_task(key):
        raise HTTPException(status_code=404, detail="Aufgabe nicht gefunden")
    response = HTMLResponse(content="")
    response.headers["HX-Toast"] = json.dumps({"message": "Aufgabe gelöscht", "type": "success"})
    return response


# ════════════════════════════════════════════════════════════════════════════
# HTML Routes — Contexts
# ════════════════════════════════════════════════════════════════════════════

def _contexts_list() -> list[dict]:
    df = list_contexts_df(include_body=False)
    if df.empty:
        return []
    return [
        {
            "key": row["Key"],
            "title": row["Titel"],
            "short_title": row.get("KurzTitel", ""),
            "short_code": row.get("Kürzel", ""),
            "description": row.get("Beschreibung", ""),
        }
        for _, row in df.iterrows()
    ]


@app.get("/contexts", response_class=HTMLResponse)
async def contexts_page(request: Request):
    return templates.TemplateResponse("contexts/index.html", {
        "request": request,
        "contexts": _contexts_list(),
    })


@app.get("/contexts/new", response_class=HTMLResponse)
async def contexts_new_form(request: Request):
    return templates.TemplateResponse("contexts/_form.html", {
        "request": request,
        "context": None,
        "body_text": "",
    })


@app.get("/contexts/{key}/edit", response_class=HTMLResponse)
async def contexts_edit_form(request: Request, key: str):
    ctx_obj, body = load_context(key)
    if not ctx_obj:
        raise HTTPException(status_code=404, detail="Kontext nicht gefunden")
    ctx = {
        "key": ctx_obj.key,
        "title": ctx_obj.title,
        "short_title": ctx_obj.short_title or "",
        "short_code": ctx_obj.short_code or "",
        "description": ctx_obj.description or "",
    }
    return templates.TemplateResponse("contexts/_form.html", {
        "request": request,
        "context": ctx,
        "body_text": body,
    })


@app.get("/contexts/{key}/confirm-delete", response_class=HTMLResponse)
async def contexts_confirm_delete(request: Request, key: str):
    ctx_obj, _ = load_context(key)
    if not ctx_obj:
        raise HTTPException(status_code=404, detail="Kontext nicht gefunden")
    return templates.TemplateResponse("contexts/_confirm_delete.html", {
        "request": request,
        "context": {"key": ctx_obj.key, "title": ctx_obj.title},
    })


@app.post("/contexts", response_class=HTMLResponse)
async def contexts_create(
    request: Request,
    title: str = Form(...),
    short_title: str = Form(""),
    short_code: str = Form(""),
    description: str = Form(""),
    body_text: str = Form(""),
):
    upsert_context(
        title=title,
        body_text=body_text,
        short_title=short_title or None,
        short_code=short_code or None,
        description=description or None,
    )
    response = templates.TemplateResponse("contexts/_table.html", {
        "request": request,
        "contexts": _contexts_list(),
    })
    response.headers["HX-Toast"] = json.dumps({"message": f'Kontext "{title}" erstellt', "type": "success"})
    return response


@app.put("/contexts/{key}", response_class=HTMLResponse)
async def contexts_update(
    request: Request,
    key: str,
    title: str = Form(...),
    short_title: str = Form(""),
    short_code: str = Form(""),
    description: str = Form(""),
    body_text: str = Form(""),
):
    ctx_obj, _ = load_context(key)
    if not ctx_obj:
        raise HTTPException(status_code=404, detail="Kontext nicht gefunden")
    upsert_context(
        key=key,
        title=title,
        body_text=body_text,
        short_title=short_title or None,
        short_code=short_code or None,
        description=description or None,
    )
    response = templates.TemplateResponse("contexts/_table.html", {
        "request": request,
        "contexts": _contexts_list(),
    })
    response.headers["HX-Toast"] = json.dumps({"message": f'Kontext "{title}" gespeichert', "type": "success"})
    return response


@app.delete("/contexts/{key}", response_class=HTMLResponse)
async def contexts_delete(key: str):
    if not soft_delete_context(key):
        raise HTTPException(status_code=404, detail="Kontext nicht gefunden")
    response = HTMLResponse(content="")
    response.headers["HX-Toast"] = json.dumps({"message": "Kontext gelöscht", "type": "success"})
    return response


@app.post("/contexts/ai-suggest", response_class=HTMLResponse)
async def contexts_ai_suggest(
    request: Request,
    ai_description: str = Form(""),
    context_key: str = Form(""),
    provider: str = Form("openai"),
    model: str = Form(""),
    temperature: float = Form(0.7),
):
    if not ai_description.strip():
        return HTMLResponse('<p class="text-muted" style="font-size:.8rem;padding:.5rem 0">Bitte Beschreibung eingeben.</p>')
    try:
        title, short_code, short_title, description, body_md = generate_context_details(
            provider=provider,
            description=ai_description.strip(),
            context_key=context_key or None,
            model=model or None,
            temperature=temperature,
        )
    except Exception as e:
        return HTMLResponse(f'<p style="color:var(--color-danger);font-size:.8rem">Fehler: {e}</p>')

    return HTMLResponse(f"""<script>
(function(){{
  var f = (id, val) => {{ var el = document.getElementById(id); if(el) el.value = val; }};
  f('c-title', {json.dumps(title)});
  f('c-short-code', {json.dumps(short_code or "")});
  f('c-short-title', {json.dumps(short_title or "")});
  f('c-description', {json.dumps(description or "")});
  f('c-body', {json.dumps(body_md or "")});
}})();
</script>
<p style="color:var(--color-success);font-size:.8rem;padding:.25rem 0">
  <svg width="12" height="12" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24" style="display:inline;vertical-align:middle"><path d="M20 6L9 17l-5-5"/></svg>
  KI-Vorschlag eingefügt!
</p>""")


# ════════════════════════════════════════════════════════════════════════════
# HTML Routes — Chat
# ════════════════════════════════════════════════════════════════════════════

def _all_models_json() -> str:
    """Returns AVAILABLE_MODELS as JSON string for use in JS."""
    result = {p: list(m.keys()) for p, m in AVAILABLE_MODELS.items()}
    return json.dumps(result)


def _default_provider() -> str:
    avail = providers_available()
    return avail[0] if avail else "openai"


@app.get("/chat", response_class=HTMLResponse)
async def chat_page(request: Request, project_key: str | None = None, provider: str | None = None):
    projects = _projects_list()
    s = _load_settings_ctx()
    sel_project = project_key or (projects[0]["key"] if projects else "")
    providers = providers_available() or ["openai"]
    sel_provider = provider or s["provider"] or providers[0]
    # Use saved model if it exists for this provider, else first available
    saved_model = s["model"] if s["model"] in get_available_models(sel_provider) else (get_available_models(sel_provider)[0] if get_available_models(sel_provider) else "")
    session_id = ""
    if sel_project and sel_provider:
        session_id = find_latest_session_for_project(sel_provider, sel_project) or str(_uuid.uuid4())
    return templates.TemplateResponse("chat/index.html", {
        "request": request,
        "projects": projects,
        "providers": providers,
        "sel_project": sel_project,
        "sel_provider": sel_provider,
        "sel_model": saved_model,
        "temperature": s["temperature"],
        "rag_enabled": s["rag_enabled"],
        "session_id": session_id,
        "all_models_json": _all_models_json(),
        "active_page": "chat",
    })


@app.get("/chat/history", response_class=HTMLResponse)
async def chat_history(
    request: Request,
    project_key: str = "",
    provider: str = "openai",
):
    history = []
    session_id = ""
    if project_key and provider:
        session_id = find_latest_session_for_project(provider, project_key) or str(_uuid.uuid4())
        if session_id:
            history = load_history(provider, session_id)
    return templates.TemplateResponse("chat/_history.html", {
        "request": request,
        "history": history,
        "session_id": session_id,
    })


@app.post("/chat/send", response_class=HTMLResponse)
async def chat_send(
    request: Request,
    project_key: str = Form(""),
    session_id: str = Form(""),
    provider: str = Form("openai"),
    model: str = Form(""),
    temperature: float = Form(0.7),
    rag_enabled: str = Form("true"),
    message: str = Form(""),
):
    if not message.strip():
        raise HTTPException(status_code=400, detail="Leere Nachricht")
    if not session_id:
        session_id = str(_uuid.uuid4())

    use_rag = rag_enabled.lower() in ("true", "1", "on", "yes")

    # Save user message
    save_message(
        provider=provider,
        session_id=session_id,
        role="user",
        content=message.strip(),
        project_key=project_key or None,
        model_name=model or None,
        model_temperature=temperature,
    )

    # Build system prompt
    project_map = ""
    proj_body = ""
    proj_title = project_key or "Unbekanntes Projekt"
    if project_key:
        try:
            proj_obj, proj_body = load_project(project_key)
            if proj_obj:
                proj_title = proj_obj.title
                project_map = build_project_map(project_key, query=message.strip())
        except Exception:
            pass

    # RAG retrieval — distillation is handled centrally inside retrieve_relevant_chunks_hybrid
    rag_section = ""
    rag_sources: list[dict] = []
    rag_results: dict = {}
    if use_rag and project_key:
        try:
            s = _load_settings_ctx()
            rag_results = retrieve_relevant_chunks_hybrid(
                message.strip(),
                project_key=project_key,
                limit=s.get("rag_top_k", 7),
                threshold=s.get("rag_similarity_threshold", 0.45),
                enable_expansion=s.get("rag_query_expansion", False),
                enable_reranking=s.get("rag_reranking_enabled", True),
            )
            rag_results = deduplicate_results(rag_results)
            rag_text = build_rag_context_from_search(rag_results)
            if rag_text:
                rag_section = f"\n\n{rag_text}"
            docs = rag_results.get("documents", [])
            rag_sources = [
                {
                    "filename": d.get("filename", ""),
                    "similarity": d.get("similarity", 0),
                    "document_id": d.get("document_id", 0),
                }
                for d in docs[:5]
            ]
        except Exception:
            pass

    system_prompt = f"""Du bist ein erfahrener Projekt-Assistent für "{proj_title}".

PROJEKT-STRUKTUR:
{project_map or "(Keine Strukturdaten verfügbar)"}

PROJEKT-BRIEF:
{proj_body or "(Kein Projektbeschrieb vorhanden)"}
{rag_section}

Antworte prägnant, strukturiert und mit direktem Bezug zu den Projekt-Anforderungen. Antworte auf Deutsch."""

    # Build chat history for context
    prev_history = load_history(provider, session_id)
    messages_for_llm = [{"role": m["role"], "content": m["content"]} for m in prev_history]
    messages_for_llm.append({"role": "user", "content": message.strip()})

    # Call LLM
    used_model: list = []
    ai_response = None
    error_text = None
    try:
        ai_response = try_models_with_messages(
            provider=provider,
            system=system_prompt,
            messages=messages_for_llm,
            max_tokens=2000,
            temperature=temperature,
            model=model or None,
            _used_model=used_model,
        )
    except Exception as e:
        error_text = str(e)

    actual_model = used_model[0] if used_model else model

    ai_msg_id = None
    user_msg_id = None
    if ai_response:
        ai_saved = save_message(
            provider=provider,
            session_id=session_id,
            role="assistant",
            content=ai_response,
            project_key=project_key or None,
            model_name=actual_model,
            model_temperature=temperature,
            rag_sources=rag_sources or None,
        )
        ai_msg_id = ai_saved.id if ai_saved else None
    else:
        ai_response = error_text or "Keine Antwort vom KI-Modell erhalten."

    from datetime import datetime
    timestamp = datetime.now().strftime("%H:%M")

    return templates.TemplateResponse("chat/_message_pair.html", {
        "request": request,
        "user_message": message.strip(),
        "ai_response": ai_response,
        "model": actual_model or model,
        "temperature": temperature,
        "rag_sources": rag_sources,
        "timestamp": timestamp,
        "error": error_text is not None and ai_response == error_text,
        "ai_msg_id": ai_msg_id,
        "user_msg_id": user_msg_id,
    })


@app.patch("/chat/messages/{msg_id}/metadata", response_class=HTMLResponse)
async def chat_update_metadata(
    msg_id: int,
    message_type: str = Form(""),
    message_status: str = Form("ungeprüft"),
):
    update_message_metadata(
        msg_id,
        message_type=message_type or None,
        message_status=message_status or "ungeprüft",
    )
    return HTMLResponse(f"""
<span style="font-size:.72rem;color:var(--color-success);padding:.1rem .3rem">
  <svg width="11" height="11" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M20 6L9 17l-5-5"/></svg>
  Gespeichert
</span>""")


@app.delete("/chat/messages/{msg_id}", response_class=HTMLResponse)
async def chat_delete_message(msg_id: int):
    delete_message(msg_id)
    return HTMLResponse("")   # OOB remove handled in template


@app.post("/chat/session/hide", response_class=HTMLResponse)
async def chat_session_hide(
    provider: str = Form(""),
    session_id: str = Form(""),
):
    if provider and session_id:
        delete_history(provider, session_id)
    response = HTMLResponse('<p class="text-muted" style="text-align:center;padding:2rem">Verlauf ausgeblendet.</p>')
    return response


@app.post("/chat/session/purge", response_class=HTMLResponse)
async def chat_session_purge(
    provider: str = Form(""),
    session_id: str = Form(""),
):
    if provider and session_id:
        purge_history(provider, session_id)
    response = HTMLResponse('<p class="text-muted" style="text-align:center;padding:2rem">Verlauf endgültig gelöscht.</p>')
    return response


@app.post("/chat/feedback", response_class=HTMLResponse)
async def chat_feedback(
    message_id: int = Form(...),
    document_id: int = Form(0),
    helpful: str = Form("true"),
):
    doc_id = document_id if document_id else 0
    is_helpful = helpful.lower() in ("true", "1", "yes")
    save_rag_feedback(message_id, doc_id, is_helpful)
    icon = "👍" if is_helpful else "👎"
    return HTMLResponse(f'<span style="font-size:.85rem">{icon}</span>')


# ════════════════════════════════════════════════════════════════════════════
# HTML Routes — KI-Vorschlag (Roles + Projects + Contexts)
# ════════════════════════════════════════════════════════════════════════════

@app.post("/roles/ai-suggest", response_class=HTMLResponse)
async def roles_ai_suggest(
    request: Request,
    ai_description: str = Form(""),
    role_key: str = Form(""),
    provider: str = Form("openai"),
    model: str = Form(""),
    temperature: float = Form(0.7),
):
    if not ai_description.strip():
        return HTMLResponse('<p class="text-muted" style="font-size:.8rem;padding:.5rem 0">Bitte Beschreibung eingeben.</p>')
    try:
        title, short_code, responsibilities, qualifications, expertise, prose = generate_role_details(
            provider=provider,
            description=ai_description.strip(),
            role_key=role_key or None,
            model=model or None,
            temperature=temperature,
        )
    except Exception as e:
        return HTMLResponse(f'<p style="color:var(--color-danger);font-size:.8rem">Fehler: {e}</p>')

    # Return JS that fills the form fields
    return HTMLResponse(f"""<script>
(function(){{
  var s = v => v.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  var f = (id, val) => {{ var el = document.getElementById(id); if(el) el.value = val; }};
  f('r-title', {json.dumps(title)});
  f('r-short-code', {json.dumps(short_code or "")});
  f('r-responsibilities', {json.dumps(responsibilities or "")});
  f('r-qualifications', {json.dumps(qualifications or "")});
  f('r-expertise', {json.dumps(expertise or "")});
  f('r-body', {json.dumps(prose or "")});
}})();
</script>
<p style="color:var(--color-success);font-size:.8rem;padding:.25rem 0">
  <svg width="12" height="12" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24" style="display:inline;vertical-align:middle"><path d="M20 6L9 17l-5-5"/></svg>
  KI-Vorschlag eingefügt!
</p>""")


@app.post("/projects/ai-suggest", response_class=HTMLResponse)
async def projects_ai_suggest(
    request: Request,
    ai_description: str = Form(""),
    project_key: str = Form(""),
    provider: str = Form("openai"),
    model: str = Form(""),
    temperature: float = Form(0.7),
):
    if not ai_description.strip():
        return HTMLResponse('<p class="text-muted" style="font-size:.8rem;padding:.5rem 0">Bitte Beschreibung eingeben.</p>')
    try:
        title, short_code, short_title, description, body_md = generate_project_details(
            provider=provider,
            description=ai_description.strip(),
            project_key=project_key or None,
            model=model or None,
            temperature=temperature,
        )
    except Exception as e:
        return HTMLResponse(f'<p style="color:var(--color-danger);font-size:.8rem">Fehler: {e}</p>')

    return HTMLResponse(f"""<script>
(function(){{
  var f = (id, val) => {{ var el = document.getElementById(id); if(el) el.value = val; }};
  f('f-title', {json.dumps(title)});
  f('f-short-code', {json.dumps(short_code or "")});
  f('f-short-title', {json.dumps(short_title or "")});
  f('f-description', {json.dumps(description or "")});
  f('f-body', {json.dumps(body_md or "")});
}})();
</script>
<p style="color:var(--color-success);font-size:.8rem;padding:.25rem 0">
  <svg width="12" height="12" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24" style="display:inline;vertical-align:middle"><path d="M20 6L9 17l-5-5"/></svg>
  KI-Vorschlag eingefügt!
</p>""")


# ════════════════════════════════════════════════════════════════════════════
# HTML Routes — Task-Generierung
# ════════════════════════════════════════════════════════════════════════════

@app.get("/taskgen", response_class=HTMLResponse)
async def taskgen_page(request: Request, role_key: str | None = None):
    roles = _roles_list()
    s = _load_settings_ctx()
    sel_role = None
    if role_key:
        for r in roles:
            if r["key"] == role_key:
                sel_role = r
                break
    import json as _json
    # Use load_role() per role to get full (non-truncated) responsibilities
    roles_full = []
    for r in roles:
        role_obj, _ = load_role(r["key"])
        roles_full.append({
            "key": r["key"],
            "title": r.get("title", ""),
            "short_code": r.get("short_code", "") or "",
            "description": r.get("description", "") or "",
            "responsibilities": (role_obj.responsibilities or "") if role_obj else (r.get("responsibilities", "") or ""),
        })
    roles_json = _json.dumps(roles_full, ensure_ascii=False)
    return templates.TemplateResponse("taskgen/index.html", {
        "request": request,
        "active_page": "taskgen",
        "roles": roles,
        "roles_json": roles_json,
        "sel_role": sel_role,
        "providers": providers_available() or ["openai"],
        "all_models_json": _all_models_json(),
        "settings": s,
    })


@app.post("/taskgen/generate", response_class=HTMLResponse)
async def taskgen_generate(
    request: Request,
    role_key: str = Form(""),
    provider: str = Form("openai"),
    model: str = Form(""),
    temperature: float = Form(0.7),
    min_per_resp: int = Form(2),
    max_per_resp: int = Form(5),
    rag_enabled: str = Form("false"),
):
    if not role_key:
        return HTMLResponse('<p class="text-muted" style="padding:1rem">Bitte eine Rolle auswählen.</p>')

    role_obj, _ = load_role(role_key)
    if not role_obj:
        return HTMLResponse('<p style="color:var(--color-danger);padding:1rem">Rolle nicht gefunden.</p>')

    responsibilities = getattr(role_obj, "responsibilities", "") or ""
    if not responsibilities.strip():
        return HTMLResponse('<p style="color:var(--color-danger);padding:1rem">Die Rolle hat keine Verantwortlichkeiten definiert. Bitte zuerst in der Rollenverwaltung ergänzen.</p>')

    s = _load_settings_ctx()
    use_rag = rag_enabled.lower() in ("true", "1", "on", "yes")

    try:
        from src.m12_task_generation import generate_tasks_from_role
        tasks = generate_tasks_from_role(
            provider=provider,
            role_title=role_obj.title,
            role_key=role_key,
            responsibilities=responsibilities,
            min_per_resp=min_per_resp,
            max_per_resp=max_per_resp,
            model_name=model or s.get("model") or None,
            temperature=temperature,
            rag_enabled=use_rag,
            rag_top_k=s.get("rag_top_k", 5),
            rag_similarity_threshold=s.get("rag_similarity_threshold", 0.45),
            rag_chunk_size=s.get("rag_chunk_size", 1000),
        )
    except Exception as e:
        return HTMLResponse(f'<div style="color:var(--color-danger);padding:1rem">Fehler beim Generieren: {e}</div>')

    return templates.TemplateResponse("taskgen/_results.html", {
        "request": request,
        "tasks": tasks,
        "role_key": role_key,
        "role_title": role_obj.title,
        "count": len(tasks),
    })


@app.post("/taskgen/save", response_class=HTMLResponse)
async def taskgen_save(request: Request):
    """Save selected generated tasks. Accepts JSON body with list of task dicts."""
    body = await request.json()
    tasks_to_save = body.get("tasks", [])
    saved = 0
    errors = []
    for t in tasks_to_save:
        try:
            upsert_task(
                title=t.get("title", ""),
                body_text=t.get("description", ""),
                short_title=t.get("title", "")[:50],
                short_code=t.get("short_code") or None,
                description=t.get("description", ""),
                source_role_key=t.get("source_role_key") or None,
                source_responsibility=t.get("source_responsibility") or None,
                generation_batch_id=t.get("generation_batch_id") or None,
            )
            saved += 1
        except Exception as e:
            errors.append(str(e))

    msg = f"{saved} Aufgabe(n) gespeichert."
    if errors:
        msg += f" {len(errors)} Fehler."
    return HTMLResponse(f"""
<div style="padding:.75rem 1rem;background:var(--color-surface-2);border:1px solid var(--color-border);border-radius:.5rem;display:flex;align-items:center;gap:.6rem">
  <svg width="16" height="16" fill="none" stroke="var(--color-success)" stroke-width="2" viewBox="0 0 24 24"><path d="M20 6L9 17l-5-5"/></svg>
  <span style="font-size:.9rem">{msg} <a href="/tasks" style="color:var(--color-primary)">Zu den Aufgaben →</a></span>
</div>""")


# ════════════════════════════════════════════════════════════════════════════
# HTML Routes — KI-Einstellungen
# ════════════════════════════════════════════════════════════════════════════

def _load_settings_ctx() -> dict:
    """Merge config defaults with user_settings.yaml overrides."""
    from src.m01_config import get_settings as _gs
    defaults = _gs().llm_defaults
    user = load_user_settings()
    merged = {**defaults, **user}
    return {
        "provider": merged.get("provider", "openai"),
        "model": merged.get("model", "gpt-4o"),
        "temperature": float(merged.get("temperature", 0.7)),
        "rag_top_k": int(merged.get("rag_top_k", 5)),
        "rag_chunk_size": int(merged.get("rag_chunk_size", 1000)),
        "rag_similarity_threshold": float(merged.get("rag_similarity_threshold", 0.45)),
        "rag_enabled": bool(merged.get("rag_enabled", True)),
        "rag_query_expansion": bool(merged.get("rag_query_expansion", False)),
        "rag_reranking_enabled": bool(merged.get("rag_reranking_enabled", True)),  # Reranking Toggle
    }


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    ctx = _load_settings_ctx()
    return templates.TemplateResponse("settings/index.html", {
        "request": request,
        "active_page": "settings",
        "all_models_json": _all_models_json(),
        "providers": ["openai", "anthropic", "mistral"],
        "provider_status": {p: have_key(p) for p in ["openai", "anthropic", "mistral"]},
        **ctx,
    })


@app.post("/settings", response_class=HTMLResponse)
async def settings_save(
    request: Request,
    provider: str = Form("openai"),
    model: str = Form(""),
    temperature: float = Form(0.7),
    rag_top_k: int = Form(5),
    rag_chunk_size: int = Form(1000),
    rag_similarity_threshold: float = Form(0.45),
    rag_enabled: str = Form("false"),
    enable_expansion: str = Form("false"),
    enable_reranking: str = Form("true"),  # Reranking Toggle
):
    settings_dict = {
        "provider": provider,
        "model": model or DEFAULT_MODELS.get(provider, ""),
        "temperature": round(temperature, 2),
        "rag_top_k": rag_top_k,
        "rag_chunk_size": rag_chunk_size,
        "rag_similarity_threshold": round(rag_similarity_threshold, 3),
        "rag_enabled": rag_enabled in ("true", "on", "1", "yes"),
        "rag_query_expansion": enable_expansion in ("true", "on", "1", "yes"),
        "rag_reranking_enabled": enable_reranking in ("true", "on", "1", "yes"),
    }
    save_user_settings(settings_dict)
    response = templates.TemplateResponse("settings/index.html", {
        "request": request,
        "active_page": "settings",
        "all_models_json": _all_models_json(),
        "providers": ["openai", "anthropic", "mistral"],
        "provider_status": {p: have_key(p) for p in ["openai", "anthropic", "mistral"]},
        **settings_dict,
        "saved": True,
    })
    response.headers["HX-Toast"] = json.dumps({"message": "Einstellungen gespeichert", "type": "success"})
    return response


@app.get("/settings/test-connection", response_class=HTMLResponse)
async def settings_test_connection(request: Request, provider: str = "openai"):
    ok, msg = test_connection(provider)
    status = "success" if ok else "error"
    icon = '<path d="M20 6L9 17l-5-5"/>' if ok else '<line x1="18" y1="6" x2="6" y2="18"/><line x1="6" y1="6" x2="18" y2="18"/>'
    color = "var(--color-success)" if ok else "var(--color-danger)"
    label = f"Verbindung OK — {provider}" if ok else f"Fehler: {msg}"
    return HTMLResponse(f"""
<span style="color:{color};font-size:.82rem;display:flex;align-items:center;gap:.35rem">
  <svg width="13" height="13" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24">{icon}</svg>
  {label}
</span>""")


# ════════════════════════════════════════════════════════════════════════════
# HTML Routes — Batch QA
# ════════════════════════════════════════════════════════════════════════════

from src.m09_docs import get_project_documents
from src.m13_ki_detector import analyze_all_vendors, analyze_vendor_with_ai
from src.m03_db import DocumentChunk, Document
from src.m07_projects import get_project_roles
from src.m09_rag import get_all_documents_with_best_scores, format_chunk_preview
from fastapi.responses import StreamingResponse
from sqlmodel import select as _select_bqa
import asyncio as _asyncio
import pathlib as _pathlib


def _strip_contextual_prefix(chunk_text: str) -> str:
    """Remove [CSV | filename | Frage X]\n prefix from chunk text."""
    if chunk_text.startswith("["):
        nl = chunk_text.find("\n")
        if nl > 0:
            return chunk_text[nl + 1:]
    return chunk_text


def _get_csv_field(data: dict, field_name: str, fallback: str = "") -> str:
    """Robust field lookup with aliases and case variants."""
    variants = [field_name, field_name.lower(), field_name.upper(),
                field_name + ".", field_name.lower() + "."]
    aliases = {
        "Nr": ["Nummer", "nummer", "Number", "number", "No", "no", "ID", "id"],
        "Frage": ["Question", "question", "Text", "text", "Frage.", "Frage?"],
        "Lieferant": ["Anbieter", "anbieter", "Vendor", "vendor", "Supplier", "supplier"],
    }
    if field_name in aliases:
        variants.extend(aliases[field_name])
    for v in variants:
        if v in data:
            val = data[v]
            return str(val) if val is not None else fallback
    return fallback


def _style_text(answer_style: str, answer_stance: str, answer_wording: str) -> str:
    style_map = {
        "Sachlich & präzise": "Antworte sachlich, präzise und auf den Punkt.",
        "Ausführlich & erkärend": "Antworte ausführlich und erkläre alle relevanten Details.",
        "Kurz & bündig": "Antworte so kurz wie möglich, max. 2-3 Sätze.",
    }
    stance_map = {
        "(nur gemäss Rolle)": "",
        "Neutral": " Bei Interpretationsspielraum wäge objektiv ab.",
        "Wohlwollend (erlaubend)": " Bei Interpretationsspielraum entscheide zugunsten des Anbieters.",
        "Restriktiv (ablehnend)": ' Bei Interpretationsspielraum entscheide ablehnend: "ist nicht zulässig", "gilt als unzulässige Subbeauftragung", "wird abgelehnt".',
    }
    wording_map = {
        "(nur gemäss Rolle)": "",
        "Klar & abschliessend": ' Formuliere verbindlich: "ist", "gilt als", "ist zulässig/unzulässig". Keine Weichmacher.',
        "Vage & mit Weichmachern": ' Formuliere offen: "sollte", "könnte", mit Klärungsvorbehalt.',
    }
    return style_map.get(answer_style, "") + stance_map.get(answer_stance, "") + wording_map.get(answer_wording, "")





def _load_bqa_questions(csv_doc_id: int) -> list[dict]:
    """Load and parse all question-rows from a CSV DocumentChunk."""
    rows = []
    with get_session() as session:
        chunks = session.exec(
            _select_bqa(DocumentChunk)
            .where(DocumentChunk.document_id == csv_doc_id)
            .order_by(DocumentChunk.chunk_index)
        ).all()
    for c in chunks:
        try:
            rows.append(json.loads(_strip_contextual_prefix(c.chunk_text)))
        except Exception:
            pass
    return rows


def _select_questions(questions: list[dict], selection_str: str) -> list[tuple[int, dict]]:
    """
    Select questions from list based on selection string.
    
    The selection is Nr-based (matches the "Nr." column in CSV) when all
    questions have a numeric Nr field. Otherwise falls back to positional
    selection (row index in DB order).
    
    Args:
        questions: List of question dicts
        selection_str: "all", "" or range like "1-10,25,51-60"
    
    Returns:
        List of (original_db_index, question_dict) tuples sorted by Nr (or position)
    """
    if not questions:
        return []
    
    sel = (selection_str or "").strip().lower()
    if sel in ("", "all"):
        return list(enumerate(questions))
    
    # Try Nr-based selection first (matches CSV Nr. column)
    nr_to_idx: dict[int, int] = {}
    for idx, q in enumerate(questions):
        nr_str = _get_csv_field(q, "Nr", "")
        try:
            nr_to_idx[int(float(nr_str))] = idx
        except (ValueError, TypeError):
            pass
    
    use_nr_based = len(nr_to_idx) == len(questions)
    
    selected_db_indices: set[int] = set()
    parts = [p.strip() for p in selection_str.split(",")]
    
    for part in parts:
        if "-" in part:
            try:
                s_str, e_str = part.split("-", 1)
                s, e = int(s_str.strip()), int(e_str.strip())
                if s > e:
                    raise ValueError(f"Bereich {s}-{e}: Start > Ende")
                if use_nr_based:
                    for n in range(s, e + 1):
                        if n in nr_to_idx:
                            selected_db_indices.add(nr_to_idx[n])
                else:
                    if s < 1 or e > len(questions):
                        raise ValueError(f"Bereich {s}-{e} ausserhalb von 1-{len(questions)}")
                    for i in range(s - 1, e):
                        selected_db_indices.add(i)
            except ValueError as exc:
                raise ValueError(str(exc)) from None
        else:
            try:
                n = int(part.strip())
                if use_nr_based:
                    if n in nr_to_idx:
                        selected_db_indices.add(nr_to_idx[n])
                else:
                    if n < 1 or n > len(questions):
                        raise ValueError(f"Frage {n} ausserhalb von 1-{len(questions)}")
                    selected_db_indices.add(n - 1)
            except ValueError as exc:
                raise ValueError(str(exc)) from None
    
    if not selected_db_indices:
        raise ValueError("Keine Fragen in Auswahl gefunden")
    
    # Sort by Nr if available, otherwise by db index
    if use_nr_based:
        def _sort_key(t: tuple[int, dict]) -> int:
            try:
                return int(float(_get_csv_field(t[1], "Nr", str(t[0] + 1)) or t[0] + 1))
            except (ValueError, TypeError):
                return t[0]
        return sorted([(idx, questions[idx]) for idx in selected_db_indices], key=_sort_key)
    else:
        return sorted([(idx, questions[idx]) for idx in selected_db_indices], key=lambda t: t[0])


def _parse_question_selection(selection_str: str, total_questions: int) -> set[int]:
    """
    Parse question selection string to 0-based index set.
    
    Syntax:
    - "all" or "" → all questions
    - "1-20" → questions 1 through 20 (inclusive)
    - "1-20,25,51-105" → ranges and individual questions
    
    Returns:
        Set of 0-based indices
    
    Raises:
        ValueError: if syntax invalid or out of bounds
    """
    if not selection_str or selection_str.lower() == "all":
        return set(range(total_questions))
    
    selected_indices: set[int] = set()
    parts = [p.strip() for p in selection_str.split(",")]
    
    for part in parts:
        if "-" in part:
            # Range
            try:
                start_str, end_str = part.split("-", 1)
                start = int(start_str.strip())
                end = int(end_str.strip())
            except ValueError:
                raise ValueError(f"Ungültige Bereichs-Syntax: '{part}' (erwartet: 'Start-Ende')")
            
            if start < 1 or end > total_questions:
                raise ValueError(f"Bereich {start}-{end} außerhalb gültiger Fragen (1-{total_questions})")
            if start > end:
                raise ValueError(f"Ungültiger Bereich: {start}-{end} (Start > Ende)")
            
            # Add all indices in range (convert 1-based to 0-based)
            for i in range(start - 1, end):
                selected_indices.add(i)
        else:
            # Single number
            try:
                num = int(part.strip())
            except ValueError:
                raise ValueError(f"Ungültige Fragennummer: '{part}' (muss Zahl sein)")
            
            if num < 1 or num > total_questions:
                raise ValueError(f"Frage {num} außerhalb gültiger Fragen (1-{total_questions})")
            
            selected_indices.add(num - 1)  # Convert to 0-based
    
    return selected_indices


@app.get("/batch-qa", response_class=HTMLResponse)
async def batch_qa_page(request: Request):
    s = _load_settings_ctx()
    projects = _projects_list()
    return templates.TemplateResponse("batch_qa/index.html", {
        "request": request,
        "active_page": "batch-qa",
        "projects": projects,
        "settings": s,
    })


@app.post("/batch-qa/csv-docs", response_class=HTMLResponse)
async def batch_qa_csv_docs(request: Request, project_key: str = Form("")):
    """HTMX: returns CSV-doc <option> tags for the selected project."""
    csv_docs = []
    if project_key:
        try:
            csv_docs = [d for d in get_project_documents(project_key)
                        if d.filename.lower().endswith(".csv")]
        except Exception:
            pass
    return templates.TemplateResponse("batch_qa/_csv_docs.html", {
        "request": request, "csv_docs": csv_docs,
    })


@app.post("/batch-qa/roles", response_class=HTMLResponse)
async def batch_qa_roles_partial(request: Request, project_key: str = Form("")):
    """HTMX: returns roles for the selected project."""
    roles = []
    if project_key:
        try:
            roles = get_project_roles(project_key) or []
        except Exception:
            pass
    return templates.TemplateResponse("batch_qa/_roles.html", {
        "request": request, "roles": roles,
    })


@app.post("/batch-qa/ki-analyze", response_class=HTMLResponse)
async def batch_qa_ki_analyze(request: Request, csv_doc_id: int = Form(0)):
    """Heuristic KI analysis on all questions in a CSV document."""
    if not csv_doc_id:
        return HTMLResponse('<p class="text-muted small">Keine CSV ausgewählt.</p>')
    try:
        questions = _load_bqa_questions(csv_doc_id)
    except Exception as e:
        return HTMLResponse(f'<div class="alert-error text-sm">DB-Fehler: {e}</div>')
    if not questions:
        return HTMLResponse('<p class="text-muted small">Keine Fragen in der CSV gefunden.</p>')
    try:
        ki_result = analyze_all_vendors(questions)
    except Exception as e:
        return HTMLResponse(f'<div class="alert-error text-sm">KI-Analyse Fehler: {e}</div>')
    return templates.TemplateResponse("batch_qa/_ki_analysis.html", {
        "request": request,
        "ki_result": ki_result,
        "vendor_count": len(ki_result.get("ranking", [])),
        "csv_doc_id": csv_doc_id,
    })


@app.post("/batch-qa/ki-ai-analyze", response_class=HTMLResponse)
async def batch_qa_ki_ai_analyze(
    request: Request,
    csv_doc_id: int = Form(0),
    vendor: str = Form(""),
    provider: str = Form("openai"),
    model: str = Form(""),
):
    """AI deep analysis for one vendor."""
    try:
        all_rows = _load_bqa_questions(csv_doc_id)
    except Exception as e:
        return HTMLResponse(f'<span class="text-danger">DB-Fehler: {e}</span>')
    vendor_qs = [
        str(q.get("Frage", "")).strip()
        for q in all_rows
        if str(q.get("Lieferant", "")).strip() == vendor and str(q.get("Frage", "")).strip()
    ]
    if not vendor_qs:
        return HTMLResponse(f'<span class="text-danger">Keine Fragen für „{vendor}" gefunden.</span>')
    try:
        ai_res = analyze_vendor_with_ai(
            questions=vendor_qs, vendor=vendor,
            provider=provider, model=model or None, temperature=0.2,
        )
    except Exception as e:
        return HTMLResponse(f'<span class="text-danger">AI-Analyse Fehler: {e}</span>')
    return templates.TemplateResponse("batch_qa/_ki_ai_result.html", {
        "request": request, "vendor": vendor, "ai_res": ai_res,
    })


@app.post("/batch-qa/prompt-preview", response_class=HTMLResponse)
async def batch_qa_prompt_preview(
    request: Request,
    project_key: str = Form(""),
    csv_doc_id: int = Form(0),
    frage_nr: int = Form(1),
    role_mode: str = Form("all_merged"),
    role_keys: str = Form(""),
    use_project_context: bool = Form(True),
    answer_style: str = Form("Sachlich & präzise"),
    answer_stance: str = Form("(nur gemäss Rolle)"),
    answer_wording: str = Form("Klar & abschliessend"),
):
    s = _load_settings_ctx()
    style_txt = _style_text(answer_style, answer_stance, answer_wording)

    # Load specific question
    error = None
    question_data = None
    try:
        rows = _load_bqa_questions(csv_doc_id)
        chunks_by_nr: dict = {}
        for cd in rows:
            nr_raw = _get_csv_field(cd, "Nr")
            try:
                nr = int(float(str(nr_raw).strip()))
            except Exception:
                nr = str(nr_raw).strip()
            chunks_by_nr[nr] = cd
        question_data = chunks_by_nr.get(frage_nr) or chunks_by_nr.get(str(frage_nr))
        if not question_data:
            available = sorted(chunks_by_nr.keys())[:20]
            error = f"Frage Nr. {frage_nr} nicht gefunden. Verfügbare Nrn: {', '.join(map(str, available))}"
    except Exception as e:
        error = str(e)

    if error:
        return HTMLResponse(f'<div style="color:var(--color-danger);padding:.75rem">{error}</div>')

    q_text = _get_csv_field(question_data, "Frage")
    q_nr = _get_csv_field(question_data, "Nr", str(frage_nr))
    q_lief = _get_csv_field(question_data, "Lieferant")

    proj_ctx = ""
    if use_project_context:
        po, _ = load_project(project_key)
        if po:
            proj_ctx = f"\n\nPROJEKT: {po.title}\nBESCHREIBUNG: {po.description or '—'}"

    # RAG
    rag_ctx = ""
    rag_sources_html = "(RAG deaktiviert)"
    pre_expansion_sources_html = ""
    rag_debug = None
    rag_diagnostics = []
    rag_warning = None
    expansion_info = {}
    reranking_info = {}
    rag_threshold = float(s.get("rag_similarity_threshold", 0.45))
    if s.get("rag_enabled", True):
        try:
            rr = retrieve_relevant_chunks_hybrid(
                q_text, project_key=project_key,
                limit=int(s.get("rag_top_k", 5)),
                threshold=rag_threshold,
                exclude_classification="FAQ/Fragen-Katalog",
                enable_expansion=s.get("rag_query_expansion", False),  # Query-Expansion nur wenn aktiviert
                enable_reranking=s.get("rag_reranking_enabled", True),  # Reranking nur wenn aktiviert
            )
            rr = deduplicate_results(rr)
            rag_debug = rr.get("debug")
            rag_ctx = build_rag_context_from_search(rr)
            rag_warning = rag_low_confidence_warning(rr, rag_threshold)
            expansion_info = rr.get("expansion", {})  # Extract expansion metadata for detailed display
            reranking_info = rr.get("reranking", {})  # Extract reranking metadata for detailed display
            docs = rr.get("documents", [])
            pre_expansion_docs = rr.get("pre_expansion_documents", [])
            # Expansion-Terme auch für Preview nutzen damit relevante Stelle angezeigt wird
            _exp = rr.get("expansion", {})
            _preview_q = (_exp.get("expansions") and q_text + " " + " ".join(_exp["expansions"].values())) or q_text
            rag_sources_html = "\n".join(
                f"• {d.get('filename','?')} ({min(max(d.get('similarity',0),d.get('normalized_match_score',0)),1.0):.0%})\n  "
                + format_chunk_preview((d.get('text','') or '').replace('\n',' '), max_length=150, query=_preview_q)
                for d in docs
            ) or "(keine passenden Chunks)"
            # 1. Lauf (vor Expansion) für Vergleich im UI
            pre_expansion_sources_html = "\n".join(
                f"• {d.get('filename','?')} ({min(max(d.get('similarity',0),d.get('normalized_match_score',0)),1.0):.0%})\n  "
                + format_chunk_preview((d.get('text','') or '').replace('\n',' '), max_length=150, query=q_text)
                for d in pre_expansion_docs
            ) if pre_expansion_docs else ""
            rag_diagnostics = get_all_documents_with_best_scores(
                (rag_debug or {}).get('keyword_query', q_text), project_key=project_key,
                threshold=rag_threshold,
                exclude_classification="FAQ/Fragen-Katalog",
            )
        except Exception:
            pass

    sel_role_keys = [k.strip() for k in role_keys.split(",") if k.strip()]
    proj_roles = get_project_roles(project_key) or []

    def _build_sys(role_obj=None, roles_list=None):
        if role_mode == "none":
            return f"Du bist ein technischer Berater.\n\n{style_txt}{proj_ctx}\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}\n\nAUFGABE: Beantworte die Frage präzise."
        elif role_mode == "all_merged":
            rc = "\n".join(f"- {r.title}: {r.description or ''}" for r in (roles_list or []))
            return f"Du bist ein technisches Berater-Team.\n\n{style_txt}{proj_ctx}\n\nTEAM-PERSPEKTIVEN:\n{rc}\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}\n\nAUFGABE: Beantworte kombiniert."
        else:
            ro = role_obj
            if ro:
                return f'Du bist Berater in der Rolle "{ro.title}".\n\nDEINE ROLLE:\n{ro.description or ""}\n\n{style_txt}{proj_ctx}\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or "(keine)"}\n\nAUFGABE: Beantworte aus deiner Rollenperspektive.'
            return "Du bist ein technischer Berater."

    relevant_roles = [r for r in proj_roles if r.key in sel_role_keys]
    sys_prompt = _build_sys(
        role_obj=relevant_roles[0] if role_mode == "single" and relevant_roles else None,
        roles_list=relevant_roles,
    )
    user_prompt = f"Frage von {q_lief} (Nr. {q_nr}):\n{q_text}"

    return templates.TemplateResponse("batch_qa/_prompt_preview.html", {
        "request": request,
        "sys_prompt": sys_prompt,
        "user_prompt": user_prompt,
        "rag_sources_html": rag_sources_html,
        "pre_expansion_sources_html": pre_expansion_sources_html,
        "rag_warning": rag_warning,
        "expansion_info": expansion_info,
        "reranking_info": reranking_info,
        "rag_diagnostics": rag_diagnostics,
        "rag_debug": rag_debug,
        "rag_threshold": rag_threshold,
        "q_nr": q_nr,
        "project_key": project_key,
        "csv_doc_id": csv_doc_id,
        "role_mode": role_mode,
        "role_keys": role_keys,
        "use_project_context": use_project_context,
    })


@app.post("/batch-qa/run-preview-llm", response_class=HTMLResponse)
async def batch_qa_run_preview_llm(
    request: Request,
    sys_prompt: str = Form(""),
    user_prompt: str = Form(""),
    provider: str = Form("openai"),
    model: str = Form(""),
    temperature: float = Form(0.7),
):
    """Single LLM call for the prompt-preview step."""
    if not sys_prompt or not user_prompt:
        return HTMLResponse('<span style="color:var(--color-danger)">Kein Prompt.</span>')
    try:
        used: list = []
        answer = try_models_with_messages(
            provider=provider, system=sys_prompt,
            messages=[{"role": "user", "content": user_prompt}],
            max_tokens=2000, temperature=temperature,
            model=model or None, _used_model=used,
        )
        used_model = used[0] if used else model
    except Exception as e:
        return HTMLResponse(f'<div style="color:var(--color-danger)">Fehler: {e}</div>')
    return templates.TemplateResponse("batch_qa/_preview_answer.html", {
        "request": request,
        "answer": answer or "(Keine Antwort)",
        "used_model": used_model,
    })


@app.post("/batch-qa/run-preview-answer", response_class=HTMLResponse)
async def batch_qa_run_preview_answer(
    request: Request,
    project_key: str = Form(""),
    csv_doc_id: int = Form(0),
    frage_nr: int = Form(1),
    role_mode: str = Form("all_merged"),
    role_keys: str = Form(""),
    use_project_context: bool = Form(True),
    answer_style: str = Form("Sachlich & präzise"),
    answer_stance: str = Form("(nur gemäss Rolle)"),
    answer_wording: str = Form("Klar & abschliessend"),
    provider: str = Form("openai"),
    model: str = Form(""),
    temperature: float = Form(0.7),
    rag_enabled: bool = Form(True),
    rag_top_k: int = Form(5),
    rag_threshold: float = Form(0.45),
):
    """
    Rebuilds the system prompt from current settings and calls the LLM.
    Always fresh — not dependent on stale textarea content.
    """
    style_txt = _style_text(answer_style, answer_stance, answer_wording)
    sel_role_keys = [k.strip() for k in role_keys.split(",") if k.strip()]

    # Load question
    error = None
    question_data = None
    try:
        rows = _load_bqa_questions(csv_doc_id)
        chunks_by_nr: dict = {}
        for cd in rows:
            nr_raw = _get_csv_field(cd, "Nr")
            try:
                nr = int(float(str(nr_raw).strip()))
            except Exception:
                nr = str(nr_raw).strip()
            chunks_by_nr[nr] = cd
        question_data = chunks_by_nr.get(frage_nr) or chunks_by_nr.get(str(frage_nr))
        if not question_data:
            error = f"Frage Nr. {frage_nr} nicht gefunden."
    except Exception as e:
        error = str(e)

    if error:
        return HTMLResponse(f'<div style="color:var(--color-danger)">{error}</div>')

    q_text = _get_csv_field(question_data, "Frage")
    q_nr = _get_csv_field(question_data, "Nr", str(frage_nr))
    q_lief = _get_csv_field(question_data, "Lieferant")

    proj_ctx = ""
    if use_project_context:
        po, _ = load_project(project_key)
        if po:
            proj_ctx = f"\n\nPROJEKT: {po.title}\nBESCHREIBUNG: {po.description or '—'}"

    # RAG
    rag_ctx = ""
    if rag_enabled:
        try:
            rr = retrieve_relevant_chunks_hybrid(
                q_text, project_key=project_key,
                limit=rag_top_k, threshold=rag_threshold,
                exclude_classification="FAQ/Fragen-Katalog",
                enable_reranking=s.get("rag_reranking_enabled", True),
            )
            rr = deduplicate_results(rr)
            rag_ctx = build_rag_context_from_search(rr)
        except Exception:
            pass

    proj_roles = get_project_roles(project_key) or []
    sel_roles = [r for r in proj_roles if r.key in sel_role_keys]

    # Build system prompt
    if role_mode == "none":
        sys_p = (f"Du bist ein technischer Berater.\n\n{style_txt}{proj_ctx}"
                 f"\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}"
                 f"\n\nAUFGABE: Beantworte die Frage präzise.")
    elif role_mode == "all_merged":
        rc = "\n".join(f"- {r.title}: {r.description or ''}" for r in sel_roles)
        sys_p = (f"Du bist ein technisches Berater-Team.\n\n{style_txt}{proj_ctx}"
                 f"\n\nTEAM-PERSPEKTIVEN:\n{rc or '(keine Rollen gewählt)'}"
                 f"\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}"
                 f"\n\nAUFGABE: Beantworte kombiniert.")
    elif role_mode == "single":
        ro = sel_roles[0] if sel_roles else None
        if ro:
            sys_p = (f'Du bist Berater in der Rolle "{ro.title}".'
                     f"\n\nDEINE ROLLE:\n{ro.description or ''}"
                     f"\n\n{style_txt}{proj_ctx}"
                     f"\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}"
                     f"\n\nAUFGABE: Beantworte aus deiner Rollenperspektive.")
        else:
            sys_p = f"Du bist ein technischer Berater.\n\n{style_txt}{proj_ctx}\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}"
    else:  # individual — first role for preview
        ro = sel_roles[0] if sel_roles else None
        if ro:
            sys_p = (f'Du bist Berater in der Rolle "{ro.title}".'
                     f"\n\nDEINE ROLLE:\n{ro.description or ''}"
                     f"\n\n{style_txt}{proj_ctx}"
                     f"\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}"
                     f"\n\nAUFGABE: Beantworte aus deiner Rollenperspektive.")
        else:
            sys_p = f"Du bist ein technischer Berater.\n\n{style_txt}{proj_ctx}\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}"

    user_p = f"Frage von {q_lief} (Nr. {q_nr}):\n{q_text}"

    try:
        used: list = []
        answer = try_models_with_messages(
            provider=provider, system=sys_p,
            messages=[{"role": "user", "content": user_p}],
            max_tokens=2000, temperature=temperature,
            model=model or None, _used_model=used,
        )
        used_model = used[0] if used else model
    except Exception as e:
        return HTMLResponse(f'<div style="color:var(--color-danger)">LLM-Fehler: {e}</div>')

    # Debug info to show settings used
    settings_summary = f"{answer_style} · {answer_stance} · {answer_wording}"
    role_summary = (
        "Keine Rolle" if role_mode == "none"
        else ("Alle: " + ", ".join(r.title for r in sel_roles) if role_mode == "all_merged"
              else (sel_roles[0].title if sel_roles else "(keine Rolle gewählt)"))
    )

    return templates.TemplateResponse("batch_qa/_preview_answer.html", {
        "request": request,
        "answer": answer or "(Keine Antwort)",
        "used_model": used_model,
        "settings_summary": settings_summary,
        "role_summary": role_summary,
    })


@app.post("/batch-qa/export-excel")
async def batch_qa_export_excel(request: Request):
    """Accept JSON body {rows: [...], filename: '...'} and return xlsx file."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")

    rows: list[dict] = body.get("rows", [])
    filename: str = body.get("filename", "batch_export") + ".xlsx"

    if not rows:
        raise HTTPException(400, "Keine Daten")

    import traceback as _traceback
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    try:
        # Collect all column names, separate display vs debug
        display_cols: list[str] = []
        debug_cols: list[str] = []
        seen: set = set()
        for row in rows:
            for k in row.keys():
                if k not in seen:
                    seen.add(k)
                    if k == "_RAG_Warning":
                        pass  # added after other display cols below
                    elif k.startswith("_"):
                        debug_cols.append(k)
                    else:
                        display_cols.append(k)
        # Always show RAG warning in main sheet (last column)
        has_warning_col = any("_RAG_Warning" in row for row in rows)
        if has_warning_col:
            display_cols.append("_RAG_Warning")

        wb = openpyxl.Workbook()

        # ── Sheet 1: Hauptergebnisse ──────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Ergebnisse"

        header_fill = PatternFill("solid", fgColor="2F5496")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        answer_fill = PatternFill("solid", fgColor="DEEAF1")
        warning_fill = PatternFill("solid", fgColor="FFF2CC")
        warning_font = Font(name="Calibri", size=10, color="7F6000")
        alt_fill = PatternFill("solid", fgColor="F2F2F2")
        normal_fill = PatternFill("solid", fgColor="FFFFFF")
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row
        for ci, col in enumerate(display_cols, 1):
            cell = ws1.cell(row=1, column=ci, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            cell.border = border

        # Data rows
        for ri, row in enumerate(rows, 2):
            fill = alt_fill if ri % 2 == 0 else normal_fill
            for ci, col in enumerate(display_cols, 1):
                val = row.get(col, "")
                if not isinstance(val, (str, int, float, bool, type(None))):
                    val = str(val)
                cell = ws1.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
                if col == "_RAG_Warning":
                    if val:
                        cell.fill = warning_fill
                        cell.font = warning_font
                    else:
                        cell.fill = normal_fill
                elif col.startswith("Antwort"):
                    cell.fill = answer_fill
                else:
                    cell.fill = fill

        # Auto column widths (capped)
        for ci, col in enumerate(display_cols, 1):
            max_w = len(col) + 2
            for row in rows:
                v = str(row.get(col, ""))
                line_len = len(v.split("\n")[0])
                max_w = max(max_w, min(line_len, 80))
            ws1.column_dimensions[get_column_letter(ci)].width = min(max_w + 2, 80)

        ws1.freeze_panes = "A2"
        ws1.row_dimensions[1].height = 22

        # ── Sheet 2: RAG Debug ────────────────────────────────────────────────
        if debug_cols:
            ws2 = wb.create_sheet("RAG Debug")
            debug_header_fill = PatternFill("solid", fgColor="375623")
            all_debug_cols = ["Nr", "Lieferant"] + debug_cols
            for ci, col in enumerate(all_debug_cols, 1):
                cell = ws2.cell(row=1, column=ci, value=col)
                cell.font = header_font
                cell.fill = debug_header_fill
                cell.alignment = Alignment(vertical="center")
                cell.border = border
            for ri, row in enumerate(rows, 2):
                for ci, col in enumerate(all_debug_cols, 1):
                    val = row.get(col, "")
                    if not isinstance(val, (str, int, float, bool, type(None))):
                        val = str(val)
                    cell = ws2.cell(row=ri, column=ci, value=val)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = border
            for ci, col in enumerate(all_debug_cols, 1):
                ws2.column_dimensions[get_column_letter(ci)].width = 12 if col in ("Nr", "Lieferant") else 60
            ws2.freeze_panes = "A2"

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from fastapi.responses import Response
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logging.error("Excel-Export fehlgeschlagen: %s\n%s", exc, _traceback.format_exc())
        raise HTTPException(500, f"Excel-Generierung fehlgeschlagen: {exc}")


@app.post("/batch-qa/checkpoint-info", response_class=HTMLResponse)
async def batch_qa_checkpoint_info(
    request: Request,
    project_key: str = Form(""),
    csv_doc_id: str = Form(""),
):
    cp = _pathlib.Path("data") / f"batch_checkpoint_{project_key}_{csv_doc_id}.json"
    if not cp.exists():
        return HTMLResponse("")
    try:
        saved = json.loads(cp.read_text(encoding="utf-8"))
        meta = saved.get("__meta__", {})
        n = len(saved.get("results", []))
        return templates.TemplateResponse("batch_qa/_checkpoint_info.html", {
            "request": request, "cp_n": n, "cp_meta": meta,
            "project_key": project_key, "csv_doc_id": csv_doc_id,
        })
    except Exception:
        return HTMLResponse("")


@app.post("/batch-qa/checkpoint-delete", response_class=HTMLResponse)
async def batch_qa_checkpoint_delete(
    project_key: str = Form(""),
    csv_doc_id: str = Form(""),
):
    cp = _pathlib.Path("data") / f"batch_checkpoint_{project_key}_{csv_doc_id}.json"
    try:
        cp.unlink(missing_ok=True)
    except Exception:
        pass
    return HTMLResponse("")


@app.get("/batch-qa/stream")
async def batch_qa_stream(
    request: Request,
    project_key: str = "",
    csv_doc_id: int = 0,
    role_mode: str = "all_merged",
    role_keys: str = "",
    use_project_context: bool = True,
    answer_style: str = "Sachlich & präzise",
    answer_stance: str = "(nur gemäss Rolle)",
    answer_wording: str = "Klar & abschliessend",
    provider: str = "openai",
    model: str = "",
    temperature: float = 0.7,
    rag_enabled: bool = True,
    rag_top_k: int = 5,
    rag_threshold: float = 0.45,
    enable_expansion: bool = False,  # Query-Expansion (Akronym-Auflösung)
    enable_reranking: bool = True,  # Reranking (Top-15→Top-7)
    question_selection: str = "all",
    force_restart: bool = False,
):
    """
    SSE stream for batch execution. Events:
      progress {current, total, nr, lieferant}
      resume   {from, total}
      result   <JSON row incl. _RAG_Chunks, _RAG_Debug>
      done     {count}
      error    {message}
    """
    style_txt = _style_text(answer_style, answer_stance, answer_wording)
    sel_role_keys = [k.strip() for k in role_keys.split(",") if k.strip()]
    current_meta = {
        "project": project_key, "csv_id": str(csv_doc_id),
        "provider": provider, "model": model,
        "role_mode": role_mode, "roles": sorted(sel_role_keys),
        "question_selection": question_selection,
    }
    cp_path = _pathlib.Path("data") / f"batch_checkpoint_{project_key}_{csv_doc_id}.json"

    async def gen():
        # ── Load questions ─────────────────────────────────────────────────
        try:
            questions = _load_bqa_questions(csv_doc_id)
        except Exception as e:
            yield f"event: error\ndata: {json.dumps({'message': str(e)})}\n\n"
            return
        if not questions:
            yield f"event: error\ndata: {json.dumps({'message': 'Keine Fragen in der CSV gefunden.'})}\n\n"
            return
        
        # ── Parse question selection ───────────────────────────────────────
        try:
            selected_questions = _select_questions(questions, question_selection)
        except ValueError as e:
            yield f"event: error\ndata: {json.dumps({'message': f'Ungültige Fragen-Auswahl: {e}'})}\n\n"
            return
        
        if not selected_questions:
            yield f"event: error\ndata: {json.dumps({'message': 'Keine Fragen ausgewählt.'})}\n\n"
            return

        # ── Checkpoint resume ──────────────────────────────────────────────
        results: list[dict] = []
        resume_from = 0
        if force_restart and cp_path.exists():
            # User explicitly wants to restart — delete existing checkpoint
            try:
                cp_path.unlink(missing_ok=True)
            except Exception:
                pass
        if not force_restart and cp_path.exists():
            try:
                saved = json.loads(cp_path.read_text(encoding="utf-8"))
                if isinstance(saved, dict) and "results" in saved:
                    sm = saved.get("__meta__", {})
                    if (sm.get("project") == project_key
                            and str(sm.get("csv_id")) == str(csv_doc_id)
                            and sm.get("provider") == provider
                            and sm.get("model") == model
                            and sm.get("role_mode") == role_mode
                            and sorted(sm.get("roles", [])) == sorted(sel_role_keys)
                            and sm.get("question_selection", "all") == question_selection):
                        results = saved["results"]
                        resume_from = len(results)
                        yield f"event: resume\ndata: {json.dumps({'from': resume_from, 'total': len(selected_questions)})}\n\n"
                    else:
                        # Settings changed — inform frontend
                        yield f"event: checkpoint_skipped\ndata: {json.dumps({'reason': 'Einstellungen geändert (Modell, Rollen oder Fragen-Auswahl)'})}\n\n"
            except Exception:
                pass

        # ── Project context ────────────────────────────────────────────────
        proj_ctx = ""
        if use_project_context:
            try:
                po, _ = load_project(project_key)
                if po:
                    proj_ctx = f"\n\nPROJEKT: {po.title}\nBESCHREIBUNG: {po.description or '—'}"
            except Exception:
                pass

        proj_roles = get_project_roles(project_key) or []

        # ── Per-question loop ──────────────────────────────────────────────
        for enum_idx, (original_idx, q_data) in enumerate(selected_questions):
            if enum_idx < resume_from:
                continue
            if await request.is_disconnected():
                return

            q_text = _get_csv_field(q_data, "Frage", "")
            nr = _get_csv_field(q_data, "Nr", str(original_idx + 1))
            lieferant = _get_csv_field(q_data, "Lieferant", "")

            yield f"event: progress\ndata: {json.dumps({'current': enum_idx+1, 'total': len(selected_questions), 'nr': nr, 'lieferant': lieferant})}\n\n"

            row: dict = {"Nr": nr, "Lieferant": lieferant, "Frage": q_text}

            # RAG
            rag_ctx = ""
            if rag_enabled:
                try:
                    rr = retrieve_relevant_chunks_hybrid(
                        q_text, project_key=project_key,
                        limit=rag_top_k, threshold=rag_threshold,
                        exclude_classification="FAQ/Fragen-Katalog",
                        enable_expansion=enable_expansion,  # Query-Expansion von URL-Parameter
                        enable_reranking=enable_reranking,  # Reranking von URL-Parameter
                    )
                    rr = deduplicate_results(rr)
                    rag_ctx = build_rag_context_from_search(rr)
                    rag_warn = rag_low_confidence_warning(rr, rag_threshold)
                    expansion_info = rr.get("expansion", {})
                    reranking_info = rr.get("reranking", {})
                    if rag_warn:
                        row["_RAG_Warning"] = rag_warn
                    if expansion_info and expansion_info.get("triggered"):
                        # Include expansion metadata for structured display in frontend
                        row["_Expansion"] = json.dumps(expansion_info)
                    if reranking_info and reranking_info.get("enabled"):
                        # Include reranking metadata for structured display in frontend
                        row["_Reranking"] = json.dumps(reranking_info)
                    # Verwende distillierte Keywords für Preview (zeigt relevante Stelle im Chunk)
                    # Falls Expansion ausgelöst wurde, sind die Expansions-Terme auch dabei
                    _preview_query = expansion_info.get("expansions") and (
                        q_text + " " + " ".join(expansion_info["expansions"].values())
                    ) or q_text
                    row["_RAG_Chunks"] = " | ".join(
                        f"{d.get('filename','?')} ({min(max(d.get('similarity',0),d.get('match_score',0)),1.0):.0%}): "
                        + format_chunk_preview((d.get('text','') or '').replace('\n',' '), max_length=120, query=_preview_query)
                        for d in rr.get("documents", [])
                    )
                    dbg = rr.get("debug", {})
                    parts = []
                    for bk, lbl in [("keyword_candidates","kw"),("semantic_candidates","sem"),("final_candidates","final")]:
                        for d in dbg.get(bk, [])[:5]:
                            cls_short = (d.get("classification") or "")[:20]
                            terms_str = ",".join(d.get("matched_terms") or []) or "—"
                            preview = (d.get("text_preview") or d.get("text") or "")[:80].replace("\n", " ")
                            parts.append(
                                f"[{lbl}] {d.get('filename','?')} "
                                f"cls={cls_short} "
                                f"comb={d.get('combined_score',0):.3f} "
                                f"sem={d.get('similarity',0):.3f} "
                                f"kw={d.get('normalized_match_score',0):.3f} "
                                f"raw={d.get('raw_bm25_score',0):.3f} "
                                f"idf={d.get('keyword_idf_score',0):.3f} "
                                f"cov={d.get('keyword_coverage',0):.3f} "
                                f"hits={d.get('priority_hits',0)} "
                                f"terms=[{terms_str}] "
                                f"preview=[{preview}]"
                            )
                    row["_RAG_Debug"] = " | ".join(parts)
                except Exception as rag_err:
                    row["_RAG_Chunks"] = f"RAG-Fehler: {rag_err}"
                    row["_RAG_Debug"] = ""
            else:
                row["_RAG_Chunks"] = "(RAG deaktiviert)"
                row["_RAG_Debug"] = "(RAG deaktiviert)"

            # Build system prompt
            def _build_sys(role_obj=None, roles_list=None):
                if role_mode == "none":
                    return (f"Du bist ein technischer Berater.\n\n{style_txt}{proj_ctx}"
                            f"\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}"
                            f"\n\nAUFGABE: Beantworte die Frage präzise.")
                elif role_mode == "all_merged":
                    rc = "\n".join(f"- {r.title}: {r.description or ''}" for r in (roles_list or []))
                    return (f"Du bist ein technisches Berater-Team.\n\n{style_txt}{proj_ctx}"
                            f"\n\nTEAM-PERSPEKTIVEN:\n{rc}"
                            f"\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}"
                            f"\n\nAUFGABE: Beantworte kombiniert.")
                else:
                    if role_obj:
                        return (f'Du bist Berater in der Rolle "{role_obj.title}".'
                                f"\n\nDEINE ROLLE:\n{role_obj.description or ''}"
                                f"\n\n{style_txt}{proj_ctx}"
                                f"\n\nRELEVANTE DOKUMENTE:\n{rag_ctx or '(keine)'}"
                                f"\n\nAUFGABE: Beantworte aus deiner Rollenperspektive.")
                    return "Du bist ein technischer Berater."

            sel_roles = [r for r in proj_roles if r.key in sel_role_keys]
            user_msg = f"Frage von {lieferant} (Nr. {nr}):\n{q_text}"

            if role_mode in ("none", "all_merged"):
                sp = _build_sys(roles_list=sel_roles)
                try:
                    used: list = []
                    ans = try_models_with_messages(
                        provider=provider, system=sp,
                        messages=[{"role": "user", "content": user_msg}],
                        max_tokens=2000, temperature=temperature,
                        model=model or None, _used_model=used,
                    ) or "[Keine Antwort]"
                except Exception as e:
                    ans = f"[Fehler: {e}]"
                row["Antwort"] = ans
            else:
                if not sel_roles:
                    row["Antwort"] = "[Keine Rolle ausgewählt]"
                for ro in sel_roles:
                    sp = _build_sys(role_obj=ro)
                    try:
                        used = []
                        ans = try_models_with_messages(
                            provider=provider, system=sp,
                            messages=[{"role": "user", "content": user_msg}],
                            max_tokens=2000, temperature=temperature,
                            model=model or None, _used_model=used,
                        ) or "[Keine Antwort]"
                    except Exception as e:
                        ans = f"[Fehler: {e}]"
                    col = "Antwort" if role_mode == "single" else f"Antwort_{ro.title}"
                    row[col] = ans

            results.append(row)

            # Save checkpoint
            try:
                _pathlib.Path("data").mkdir(exist_ok=True)
                cp_path.write_text(
                    json.dumps({"__meta__": current_meta, "results": results},
                               ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
            except Exception:
                pass

            yield f"event: result\ndata: {json.dumps(row, ensure_ascii=False)}\n\n"
            await _asyncio.sleep(0)

        # Clean checkpoint on success
        try:
            cp_path.unlink(missing_ok=True)
        except Exception:
            pass

        yield f"event: done\ndata: {json.dumps({'count': len(results)})}\n\n"

    return StreamingResponse(gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ════════════════════════════════════════════════════════════════════════════
# JSON REST API (bleibt erhalten für externe Clients / Swagger)
# ════════════════════════════════════════════════════════════════════════════

# ── Pydantic models ────────────────────────────────────────────────────────

class RoleCreate(BaseModel):
    title: str
    group_name: Optional[str] = None
    body_text: Optional[str] = ""

class ProjectCreate(BaseModel):
    title: str
    description: str
    short_title: Optional[str] = None
    short_code: Optional[str] = None
    type: Optional[str] = None
    body_text: Optional[str] = ""

# ── Roles ──────────────────────────────────────────────────────────────────

@app.get("/api/roles", response_model=List[dict], tags=["Roles"])
async def get_roles():
    """Alle Rollen abrufen"""
    try:
        df = list_roles_df(include_deleted=False)
        if df is None or df.empty:
            return []
        return df.to_dict("records")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/roles/{role_key}", tags=["Roles"])
async def get_role(role_key: str):
    """Einzelne Rolle abrufen"""
    role_obj, role_body = load_role(role_key)
    if not role_obj:
        raise HTTPException(status_code=404, detail="Rolle nicht gefunden")
    return {
        "key": role_obj.key,
        "title": role_obj.title,
        "short_code": role_obj.short_code,
        "body_text": role_body or "",
    }


@app.post("/api/roles", tags=["Roles"])
async def create_role(role: RoleCreate):
    """Neue Rolle erstellen"""
    try:
        role_obj, created = upsert_role(title=role.title, body_text=role.body_text)
        return {"key": role_obj.key, "title": role_obj.title, "short_code": role_obj.short_code, "created": created}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/roles/{role_key}", tags=["Roles"])
async def delete_role(role_key: str):
    """Rolle löschen"""
    if not soft_delete_role(role_key):
        raise HTTPException(status_code=404, detail="Rolle nicht gefunden")
    return {"success": True}


# ── Projects ───────────────────────────────────────────────────────────────

@app.get("/api/projects", response_model=List[dict], tags=["Projects"])
async def api_get_projects():
    """Alle Projekte abrufen (JSON)"""
    return _projects_list()


@app.get("/api/projects/{key}", tags=["Projects"])
async def api_get_project(key: str):
    """Einzelnes Projekt abrufen (JSON)"""
    proj, body = load_project(key)
    if not proj:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    return {
        "key": proj.key, "title": proj.title, "short_title": proj.short_title,
        "short_code": proj.short_code, "type": proj.type, "description": proj.description,
        "body_text": body,
        "role_keys": json.loads(proj.role_keys) if proj.role_keys else [],
        "context_keys": json.loads(proj.context_keys) if proj.context_keys else [],
    }


@app.post("/api/projects", tags=["Projects"])
async def api_create_project(project: ProjectCreate):
    """Neues Projekt erstellen (JSON)"""
    try:
        obj, created = upsert_project(
            title=project.title,
            description=project.description,
            short_title=project.short_title,
            short_code=project.short_code,
            type_name=project.type,
            body_text=project.body_text or "",
        )
        return {"key": obj.key, "title": obj.title, "created": created}
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/projects/{key}", tags=["Projects"])
async def api_delete_project(key: str):
    """Projekt löschen (JSON)"""
    if not soft_delete_project(key):
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    return {"success": True}


# ── LLM ───────────────────────────────────────────────────────────────────

@app.get("/api/llm/providers", tags=["LLM"])
async def get_llm_providers():
    try:
        return {"providers": providers_available()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/llm/generate-role", tags=["LLM"])
async def generate_role_endpoint(provider: str, title: str, group_name: Optional[str] = None):
    try:
        return {"generated_text": generate_role_text(provider, title, group_name)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/llm/generate-summary", tags=["LLM"])
async def generate_summary_endpoint(provider: str, title: str, content: str):
    try:
        return {"summary": generate_summary(provider, title, content)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Health ─────────────────────────────────────────────────────────────────

@app.get("/health", tags=["System"])
async def health_check():
    return {"status": "healthy", "message": "SlitProjektHub API is running", "version": "2.0.0"}


# ── Jinja2 globals: inject KI settings + model list into every template ───

def _ki_global_json() -> str:
    """Returns all KI settings + providers + allModels as single JSON string for window.KI."""
    ctx = _load_settings_ctx()
    ctx["providers"] = providers_available() or ["openai"]
    ctx["allModels"] = {p: list(m.keys()) for p, m in AVAILABLE_MODELS.items()}
    ctx["keyStatus"] = {p: have_key(p) for p in ["openai", "anthropic", "mistral"]}
    return json.dumps(ctx)

templates.env.globals["_ki_global_json"] = _ki_global_json


# ── Entry point ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
