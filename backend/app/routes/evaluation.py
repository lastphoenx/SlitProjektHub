"""FastAPI routes — Phase C Offertbeurteilung."""
from __future__ import annotations

import asyncio
import csv
import io
import json

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from backend.app.jinja_env import templates

from src.m07_projects import list_projects_df
from src.m09_docs import (
    force_rechunk_document,
    get_project_documents,
    ingest_document,
    calculate_sha256,
    get_document_by_sha256,
    link_document_to_project,
)
from src.m14_auth import (
    can_evaluate,
    can_view_evaluator_details,
    get_user_id,
    is_super_user,
    session_username,
)
from src.m15_evaluation import (
    ANGEbot_CLASSIFICATION,
    ANGEbot_SUBTYPES,
    CRITERION_KINDS,
    PRICE_FORMULA_LABELS,
    PRICE_FORMULAS,
    RANKING_PHASE_LABELS,
    RANKING_PHASES,
    TENDER_ROLE_LABELS,
    TENDER_ROLES,
    compute_bidder_tco,
    build_evaluation_export_sheets,
    compute_rankings,
    create_bidder,
    create_criterion,
    delete_price_item,
    extract_criteria_from_tender_docs,
    criteria_apply_requires_confirm,
    criteria_editor_payload,
    criteria_preview_meta,
    extract_price_from_bidder_doc,
    extract_price_structure_from_tender,
    get_bidder_document_ids,
    get_bidder_doc_subtypes,
    get_bidder_preisblatt_doc_ids,
    get_evaluation_config,
    get_score,
    get_tender_document_ids,
    import_criteria_payload,
    ki_busy_hint,
    link_document_to_bidder,
    list_bidders,
    list_criteria,
    list_missing_justifications,
    list_evaluator_score_discrepancies,
    list_price_items,
    list_scores_for_cell,
    list_scores_for_project,
    list_tender_docs,
    load_criteria_preview,
    store_criteria_preview,
    merge_price_structure_for_bidder,
    normalize_chunk_size,
    official_score,
    price_offers_status,
    recommended_chunk_size,
    project_evaluation_started,
    resolve_vorgaben_ki,
    rolled_up_score,
    score_requires_justification,
    save_criteria_editor_payload,
    save_evaluation_config,
    seed_price_structure_for_bidder,
    set_tender_doc_roles,
    set_bidder_doc_subtypes,
    soft_delete_bidder,
    soft_delete_criterion,
    suggest_score_with_rag,
    suggest_tender_role,
    sync_price_criterion_scores,
    unlink_document_from_bidder,
    upsert_price_item,
    update_criterion_ranking_phase,
    upsert_score,
    validate_evaluation_cloud_gate,
    validate_criterion_change_during_evaluation,
    validate_criteria_manage_save,
    validate_tender_cloud_gate,
)
from src.m01_config import load_user_settings

router = APIRouter()


def _username(request: Request) -> str:
    from src.m14_auth import validate_session_token
    from src.m01_config import get_settings

    token = request.cookies.get("_auth_token", "")
    s = get_settings()
    if validate_session_token(token, max_age_seconds=s.auth_session_timeout_minutes * 60):
        return session_username(token) or ""
    return ""


def _projects_list() -> list[dict]:
    df = list_projects_df()
    if df is None or df.empty:
        return []
    return [{"key": row["Key"], "title": row["Titel"]} for _, row in df.iterrows()]


def _project_title(project_key: str) -> str:
    for p in _projects_list():
        if p["key"] == project_key:
            return p["title"]
    return project_key


def _llm_picker_context() -> dict:
    from src.m16_idea_visual import (
        visual_text_models_map,
        visual_text_providers_available,
        visual_vision_models_map,
    )

    settings = load_user_settings()
    return {
        "llm_provider": settings.get("provider", "openai"),
        "llm_model": settings.get("model", ""),
        "visual_llm_providers": visual_text_providers_available(),
        "visual_llm_models": visual_text_models_map(),
        "visual_vision_models": visual_vision_models_map(),
        "form_visual_provider": "",
        "form_visual_model": "",
        "form_input_provider": "",
        "form_input_model": "",
    }


def _criteria_preview_template_ctx(request: Request, project_key: str, body: dict) -> dict:
    ctx = {
        "request": request,
        "project_key": project_key,
        "project_title": _project_title(project_key),
        "may_evaluate": True,
        "ranking_phase_labels": RANKING_PHASE_LABELS,
        "ranking_phases": RANKING_PHASES,
        **body,
    }
    ctx.update(_llm_picker_context())
    return ctx


def _criteria_preview_redirect(project_key: str, **body) -> RedirectResponse:
    preview_id = store_criteria_preview(project_key, body)
    return RedirectResponse(
        url=f"/evaluation/criteria-preview?project_key={project_key}&preview_id={preview_id}",
        status_code=303,
    )


@router.get("/evaluation/ki-hint")
async def evaluation_ki_hint(provider: str = "", model: str = ""):
    """Ollama-VRAM / Warteschlange — für Busy-Banner in der Offert-UI."""
    return ki_busy_hint(provider, model)


@router.get("/evaluation", response_class=HTMLResponse)
async def evaluation_page(request: Request, project_key: str = ""):
    projects = _projects_list()
    if not project_key and projects:
        project_key = projects[0]["key"]

    bidders = list_bidders(project_key) if project_key else []
    criteria = list_criteria(project_key) if project_key else []
    scores = list_scores_for_project(project_key) if project_key else []
    scores_by_cell: dict[tuple[int, int], list] = {}
    for s in scores:
        scores_by_cell.setdefault((s.bidder_id, s.criterion_id), []).append(s)
    rankings = compute_rankings(project_key) if project_key else []

    # Nur Top-Level-Kriterien in der Matrix - Unterfragen (parent_id gesetzt) sind
    # Beleg-/KI-Hilfsebene und werden über die Zell-Details der Elternzeile erreicht.
    top_criteria = [c for c in criteria if c.parent_id is None]
    has_children = {c.parent_id for c in criteria if c.parent_id}
    matrix_rows = []
    for crit in top_criteria:
        row = {
            "id": crit.id,
            "name": crit.name,
            "kind": crit.kind,
            "auto_price": crit.auto_price,
            "has_children": crit.id in has_children,
            "weight": crit.weight_pct if crit.kind == "zuschlag" else None,
        }
        crit_children = [c for c in criteria if c.parent_id == crit.id]
        cells = []
        for bidder in bidders:
            cell_scores = scores_by_cell.get((bidder.id, crit.id), [])
            ai_row = next((s for s in cell_scores if s.source_key == "ai"), None)
            user_rows = [s for s in cell_scores if s.source_key.startswith("user:")]
            if crit.kind == "zuschlag":
                official, answered, total = rolled_up_score(bidder.id, crit, criteria, scores_by_cell)
                display = f"{official:.2f}" if official is not None else "—"
            elif crit_children:
                child_vals = [official_score(bidder.id, ch, scores_by_cell.get((bidder.id, ch.id), [])) for ch in crit_children]
                answered = sum(1 for v in child_vals if v is not None)
                total = len(crit_children)
                official = None
                if answered == total and total:
                    official = 0.0 if any(v == 0 for v in child_vals) else 1.0
                display = "K.O." if official == 0.0 else ("erfüllt" if official == 1.0 else f"{answered}/{total}")
            else:
                official = official_score(bidder.id, crit, cell_scores)
                answered, total = (1, 1) if official is not None else (0, 1)
                display = ("Ja" if official == 1 else "Nein") if official is not None else "—"
            cells.append(
                {
                    "bidder_id": bidder.id,
                    "criterion_id": crit.id,
                    "official": official,
                    "display": display,
                    "ai_value": ai_row.value if ai_row else None,
                    "evaluator_count": len(user_rows),
                    "answered": answered,
                    "total": total,
                }
            )
        row["cells"] = cells
        matrix_rows.append(row)

    offer_docs = []
    bidder_docs: dict[int, list[int]] = {}
    bidder_doc_subtypes: dict[int, dict[int, list[str]]] = {}
    tender_doc_roles: dict[int, list[str]] = {}
    tender_role_suggestions: dict[int, str | None] = {}
    chunk_recommendations: dict[int, int] = {}
    project_source_docs = []
    eval_config = get_evaluation_config(project_key) if project_key else {}
    if project_key:
        all_project_docs = get_project_documents(project_key)
        offer_docs = [
            d for d in all_project_docs
            if d.classification == ANGEbot_CLASSIFICATION
        ]
        project_source_docs = [
            d for d in all_project_docs
            if d.classification != ANGEbot_CLASSIFICATION
        ]
        for row in list_tender_docs(project_key):
            tender_doc_roles.setdefault(row.document_id, []).append(row.tender_role)
        for doc in project_source_docs:
            tender_role_suggestions[doc.id] = suggest_tender_role(doc.classification, doc.filename)
            doc_roles = tender_doc_roles.get(doc.id, [])
            chunk_recommendations[doc.id] = recommended_chunk_size(
                doc.classification,
                tender_role=doc_roles[0] if doc_roles else None,
                filename=doc.filename,
            )
        for b in bidders:
            bidder_docs[b.id] = get_bidder_document_ids(b.id)
            bidder_doc_subtypes[b.id] = {}
            for doc_id in bidder_docs[b.id]:
                bidder_doc_subtypes[b.id][doc_id] = get_bidder_doc_subtypes(b.id, doc_id)

    who = _username(request)
    user_id = get_user_id(who) if who else None
    ctx = {
        "request": request,
        "active_page": "evaluation",
        "projects": projects,
        "project_key": project_key,
        "project_title": _project_title(project_key) if project_key else "",
        "bidders": bidders,
        "criteria": criteria,
        "top_criteria": top_criteria,
        "rankings": rankings,
        "matrix_rows": matrix_rows,
        "offer_docs": offer_docs,
        "project_source_docs": project_source_docs,
        "tender_doc_roles": tender_doc_roles,
        "tender_role_suggestions": tender_role_suggestions,
        "chunk_recommendations": chunk_recommendations,
        "eval_config": eval_config,
        "tender_roles": TENDER_ROLES,
        "tender_role_labels": TENDER_ROLE_LABELS,
        "has_tender_docs": bool(get_tender_document_ids(project_key)) if project_key else False,
        "bidder_docs": bidder_docs,
        "bidder_doc_subtypes": bidder_doc_subtypes,
        "has_phase2_criteria": any(
            int(c.ranking_phase or 1) >= 2
            for c in criteria
            if c.kind == "zuschlag" and c.parent_id is None
        ) if project_key else False,
        "ranking_phase_labels": RANKING_PHASE_LABELS,
        "ranking_phases": RANKING_PHASES,
        "price_formula_labels": PRICE_FORMULA_LABELS,
        "price_formulas": PRICE_FORMULAS,
        "missing_justifications": list_missing_justifications(project_key) if project_key else [],
        "evaluator_discrepancies": list_evaluator_score_discrepancies(project_key) if project_key else [],
        "evaluation_started": project_evaluation_started(project_key) if project_key else False,
        "price_offers_status": price_offers_status(project_key) if project_key else {},
        "angebot_class": ANGEbot_CLASSIFICATION,
        "angebot_subtypes": ANGEbot_SUBTYPES,
        "may_evaluate": can_evaluate(who),
        "may_see_evaluators": can_view_evaluator_details(who),
        "super_user": is_super_user(who),
        "user_id": user_id,
        "error": None,
        "message": None,
        "doc_upload": request.query_params.get("doc_upload"),
    }
    ctx.update(_llm_picker_context())
    return templates.TemplateResponse("evaluation/index.html", ctx)


@router.post("/evaluation/bidder", response_class=HTMLResponse)
async def evaluation_create_bidder(request: Request, project_key: str = Form(...), name: str = Form(...)):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    create_bidder(project_key, name)
    return RedirectResponse(url=f"/evaluation?project_key={project_key}", status_code=303)


@router.post("/evaluation/criterion", response_class=HTMLResponse)
async def evaluation_create_criterion(
    request: Request,
    project_key: str = Form(...),
    name: str = Form(...),
    kind: str = Form(...),
    weight_pct: float = Form(0.0),
    scale_max: int = Form(10),
    parent_id: str = Form(""),
    auto_price: str = Form("false"),
    description: str = Form(""),
    ranking_phase: str = Form(""),
    confirm_active_evaluation: str = Form(""),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    confirmed = confirm_active_evaluation.lower() in ("true", "on", "1", "yes")
    try:
        validate_criterion_change_during_evaluation(
            project_key, confirm_active_evaluation=confirmed, structural=True,
        )
    except ValueError:
        return RedirectResponse(
            url=f"/evaluation?project_key={project_key}&criteria_change=confirm_required",
            status_code=303,
        )
    rp: int | None = None
    if ranking_phase.strip().isdigit():
        rp = int(ranking_phase.strip())
    create_criterion(
        project_key,
        kind,
        name,
        weight_pct=weight_pct,
        scale_max=scale_max,
        parent_id=int(parent_id) if parent_id.strip() else None,
        auto_price=auto_price in ("true", "on", "1", "yes"),
        description=description or None,
        ranking_phase=rp,
    )
    return RedirectResponse(url=f"/evaluation?project_key={project_key}", status_code=303)


@router.get("/evaluation/criteria-manage", response_class=HTMLResponse)
async def evaluation_criteria_manage_page(request: Request, project_key: str):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    if not project_key:
        return RedirectResponse(url="/evaluation", status_code=303)
    payload = criteria_editor_payload(project_key)
    return templates.TemplateResponse(
        "evaluation/_criteria_manage.html",
        {
            "request": request,
            "project_key": project_key,
            "project_title": _project_title(project_key),
            "payload": payload,
            "ranking_phase_labels": RANKING_PHASE_LABELS,
            "ranking_phases": RANKING_PHASES,
            "evaluation_started": project_evaluation_started(project_key),
        },
    )


@router.post("/evaluation/criteria-save", response_class=HTMLResponse)
async def evaluation_criteria_save(
    request: Request,
    project_key: str = Form(...),
    criteria_json: str = Form(...),
    deleted_criterion_ids: str = Form("[]"),
    confirm_active_evaluation: str = Form(""),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    try:
        data = json.loads(criteria_json or "{}")
    except json.JSONDecodeError:
        return RedirectResponse(
            url=f"/evaluation/criteria-manage?project_key={project_key}&error=invalid",
            status_code=303,
        )
    try:
        deleted = json.loads(deleted_criterion_ids or "[]")
        if not isinstance(deleted, list):
            deleted = []
    except json.JSONDecodeError:
        deleted = []
    confirmed = confirm_active_evaluation.lower() in ("true", "on", "1", "yes")
    try:
        save_criteria_editor_payload(
            project_key, data, deleted_ids=deleted, confirm_active_evaluation=confirmed,
        )
    except ValueError:
        return RedirectResponse(
            url=f"/evaluation/criteria-manage?project_key={project_key}&error=confirm_required",
            status_code=303,
        )
    return RedirectResponse(
        url=f"/evaluation/criteria-manage?project_key={project_key}&saved=1",
        status_code=303,
    )


@router.post("/evaluation/criterion-phase", response_class=HTMLResponse)
async def evaluation_criterion_phase(
    request: Request,
    project_key: str = Form(...),
    criterion_id: int = Form(...),
    ranking_phase: int = Form(1),
    confirm_active_evaluation: str = Form(""),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    confirmed = confirm_active_evaluation.lower() in ("true", "on", "1", "yes")
    try:
        validate_criterion_change_during_evaluation(
            project_key, confirm_active_evaluation=confirmed, structural=True,
        )
    except ValueError:
        return RedirectResponse(
            url=f"/evaluation?project_key={project_key}&criteria_change=confirm_required",
            status_code=303,
        )
    try:
        update_criterion_ranking_phase(criterion_id, ranking_phase)
    except ValueError:
        return RedirectResponse(
            url=f"/evaluation?project_key={project_key}&criterion_phase=error",
            status_code=303,
        )
    return RedirectResponse(url=f"/evaluation?project_key={project_key}", status_code=303)


@router.post("/evaluation/score", response_class=HTMLResponse)
async def evaluation_save_score(
    request: Request,
    project_key: str = Form(...),
    bidder_id: int = Form(...),
    criterion_id: int = Form(...),
    value: float = Form(...),
    justification: str = Form(""),
    source_chunk_ref: str = Form(""),
    target_user_id: str = Form(""),
):
    who = _username(request)
    if not can_evaluate(who):
        raise HTTPException(403, "Keine Berechtigung")
    uid = get_user_id(who)
    if not uid:
        raise HTTPException(401, "Nicht angemeldet")
    # Standard: eigene Zeile. Nur Super-User darf eine fremde (target_user_id) korrigieren.
    write_uid = uid
    if target_user_id.strip():
        if not is_super_user(who):
            raise HTTPException(403, "Nur Super-User dürfen fremde Bewertungen ändern")
        write_uid = int(target_user_id)
    try:
        upsert_score(
            bidder_id,
            criterion_id,
            write_uid,
            value,
            justification=justification or None,
            source_chunk_ref=source_chunk_ref or None,
        )
    except ValueError as exc:
        if request.headers.get("hx-request"):
            return await evaluation_cell(
                request,
                bidder_id=bidder_id,
                criterion_id=criterion_id,
                project_key=project_key,
                score_error=str(exc),
            )
        raise HTTPException(400, str(exc)) from exc
    if request.headers.get("hx-request"):
        return await evaluation_cell(request, bidder_id=bidder_id, criterion_id=criterion_id, project_key=project_key)
    return RedirectResponse(url=f"/evaluation?project_key={project_key}", status_code=303)


@router.post("/evaluation/score/ai", response_class=HTMLResponse)
async def evaluation_save_ai_score(
    request: Request,
    project_key: str = Form(...),
    bidder_id: int = Form(...),
    criterion_id: int = Form(...),
    value: float = Form(...),
    justification: str = Form(""),
    source_chunk_ref: str = Form(""),
):
    """Speichert einen KI-Vorschlag als eigene, klar markierte Spalte (source='ai') -
    NICHT als Bewertung einer Person. Getrennt von /evaluation/score."""
    who = _username(request)
    if not can_evaluate(who):
        raise HTTPException(403, "Keine Berechtigung")
    upsert_score(
        bidder_id,
        criterion_id,
        evaluator_user_id=0,
        value=value,
        justification=justification or None,
        source_chunk_ref=source_chunk_ref or None,
        as_source="ai",
    )
    if request.headers.get("hx-request"):
        return await evaluation_cell(request, bidder_id=bidder_id, criterion_id=criterion_id, project_key=project_key)
    return RedirectResponse(url=f"/evaluation?project_key={project_key}", status_code=303)


@router.get("/evaluation/cell", response_class=HTMLResponse)
async def evaluation_cell(
    request: Request,
    bidder_id: int,
    criterion_id: int,
    project_key: str = "",
    score_error: str = "",
):
    """Detail-Panel einer Matrix-Zelle: KI-Vorschlag + jede Bewerter-Zeile einzeln,
    plus Formular fuer die eigene Bewertung. Das ist die 'mehrere Spalten'-Ansicht."""
    from src.m15_evaluation import Criterion, Bidder
    from src.m03_db import get_session

    who = _username(request)
    with get_session() as session:
        crit = session.get(Criterion, criterion_id)
        bidder = session.get(Bidder, bidder_id)
    if not crit or not bidder:
        raise HTTPException(404, "Nicht gefunden")

    all_criteria = list_criteria(project_key) if project_key else []
    children = [c for c in all_criteria if c.parent_id == crit.id]

    if children:
        # Elternkriterium mit Einzelanforderungen: gemaess Ausschreibungsvorgabe
        # ("Punkte = erreichte Punktzahl / Anzahl der Einzelanforderungen") wird
        # NICHT das Elternkriterium direkt bewertet, sondern jede Anforderung
        # einzeln - der Mittelwert ergibt automatisch den Elternwert.
        scores_by_cell: dict[tuple[int, int], list] = {}
        for s in list_scores_for_project(project_key):
            scores_by_cell.setdefault((s.bidder_id, s.criterion_id), []).append(s)
        child_rows = []
        for ch in children:
            cell = scores_by_cell.get((bidder_id, ch.id), [])
            ai_row = next((s for s in cell if s.source_key == "ai"), None)
            child_rows.append(
                {
                    "criterion": ch,
                    "official": official_score(bidder_id, ch, cell),
                    "ai_value": ai_row.value if ai_row else None,
                    "evaluator_count": len([s for s in cell if s.source_key.startswith("user:")]),
                }
            )
        rollup_val, answered, total = rolled_up_score(bidder_id, crit, all_criteria, scores_by_cell)
        return templates.TemplateResponse(
            "evaluation/_cell_parent.html",
            {
                "request": request,
                "project_key": project_key,
                "bidder": bidder,
                "criterion": crit,
                "child_rows": child_rows,
                "official": rollup_val,
                "answered": answered,
                "total": total,
                "may_evaluate": can_evaluate(who),
            },
        )

    cell_scores = list_scores_for_cell(bidder_id, criterion_id)
    ai_row = next((s for s in cell_scores if s.source_key == "ai"), None)
    user_rows = [s for s in cell_scores if s.source_key.startswith("user:")]
    uid = get_user_id(who) if who else None
    ctx = {
        "request": request,
        "project_key": project_key,
        "bidder": bidder,
        "criterion": crit,
        "ai_row": ai_row,
        "user_rows": user_rows,
        "official": official_score(bidder_id, crit, cell_scores),
        "may_evaluate": can_evaluate(who),
        "may_see_evaluators": can_view_evaluator_details(who),
        "super_user": is_super_user(who),
        "my_user_id": uid,
        "has_bidder_docs": bool(get_bidder_document_ids(bidder_id)),
        "score_error": score_error,
        "requires_justification": score_requires_justification,
    }
    ctx.update(_llm_picker_context())
    return templates.TemplateResponse("evaluation/_cell.html", ctx)


@router.post("/evaluation/suggest", response_class=HTMLResponse)
async def evaluation_suggest(
    request: Request,
    project_key: str = Form(...),
    bidder_id: int = Form(...),
    criterion_id: int = Form(...),
    provider: str = Form("openai"),
    model: str = Form(""),
    visual_llm_provider: str = Form(""),
    visual_llm_model: str = Form(""),
    cloud_confirm: str = Form(""),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    from src.m15_evaluation import Criterion
    from src.m03_db import get_session
    from src.m16_idea_visual import resolve_visual_llm

    with get_session() as session:
        crit = session.get(Criterion, criterion_id)
    if not crit:
        raise HTTPException(404, "Kriterium nicht gefunden")

    settings = load_user_settings()
    fallback_p = (provider or "").strip() or settings.get("provider", "openai")
    fallback_m = (model or "").strip() or settings.get("model", "")
    ap, am = resolve_visual_llm(visual_llm_provider, visual_llm_model, fallback_p, fallback_m)
    gate_err = validate_evaluation_cloud_gate(
        ap, bidder_id, cloud_confirm in ("1", "true", "on", "yes")
    )
    tpl_ctx = {
        "request": request,
        "suggestion": {"value": None},
        "gate_error": gate_err,
        "cloud_provider": ap,
        "project_key": project_key,
        "bidder_id": bidder_id,
        "criterion_id": criterion_id,
    }
    if gate_err:
        return templates.TemplateResponse("evaluation/_suggestion.html", tpl_ctx, status_code=400)

    suggestion = suggest_score_with_rag(
        project_key,
        bidder_id,
        crit,
        provider=ap,
        model=am or None,
    )
    tpl_ctx["suggestion"] = suggestion
    tpl_ctx["gate_error"] = None
    return templates.TemplateResponse("evaluation/_suggestion.html", tpl_ctx)


@router.post("/evaluation/tender-doc", response_class=HTMLResponse)
async def evaluation_tender_doc(
    request: Request,
    project_key: str = Form(...),
    document_id: int = Form(...),
    tender_roles: list[str] = Form([]),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    set_tender_doc_roles(project_key, document_id, tender_roles)
    return RedirectResponse(url=f"/evaluation?project_key={project_key}", status_code=303)


@router.post("/evaluation/tender-doc-upload", response_class=HTMLResponse)
async def evaluation_tender_doc_upload(
    request: Request,
    project_key: str = Form(...),
    file: UploadFile = File(...),
    classification: str = Form(""),
    tender_roles: list[str] = Form([]),
    chunk_size: int = Form(1000),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    file_bytes = await file.read()
    redirect = f"/evaluation?project_key={project_key}"
    if not file_bytes or not file.filename:
        return RedirectResponse(url=f"{redirect}&tender_upload=error", status_code=303)

    from src.m03_db import DOCUMENT_CLASSIFICATIONS

    cls = (classification or "").strip() or "Sonstiges"
    if cls not in DOCUMENT_CLASSIFICATIONS:
        cls = "Sonstiges"
    roles = [
        (r or "").strip().lower()
        for r in tender_roles
        if (r or "").strip().lower() in TENDER_ROLES
    ]
    if not roles:
        return RedirectResponse(url=f"{redirect}&tender_upload=error", status_code=303)
    chunk_size = normalize_chunk_size(
        chunk_size,
        fallback=recommended_chunk_size(cls, tender_role=roles[0], filename=file.filename or ""),
    )

    success, _msg = ingest_document(
        file_name=file.filename,
        file_bytes=file_bytes,
        classification=cls,
        chunk_size=chunk_size,
    )
    doc = get_document_by_sha256(calculate_sha256(file_bytes), include_deleted=True)
    if not doc:
        return RedirectResponse(url=f"{redirect}&tender_upload=error", status_code=303)
    link_document_to_project(project_key, doc.id)
    set_tender_doc_roles(project_key, doc.id, roles)
    status = "ok" if success else "linked"
    return RedirectResponse(url=f"{redirect}&tender_upload={status}", status_code=303)


@router.post("/evaluation/bidder-doc-subtypes", response_class=HTMLResponse)
async def evaluation_bidder_doc_subtypes(
    request: Request,
    project_key: str = Form(...),
    bidder_id: int = Form(...),
    document_id: int = Form(...),
    doc_subtypes: list[str] = Form([]),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    try:
        set_bidder_doc_subtypes(bidder_id, document_id, doc_subtypes)
    except ValueError:
        return RedirectResponse(
            url=f"/evaluation?project_key={project_key}&doc_subtype=error",
            status_code=303,
        )
    return RedirectResponse(url=f"/evaluation?project_key={project_key}", status_code=303)


@router.post("/evaluation/bidder-doc", response_class=HTMLResponse)
async def evaluation_bidder_doc(
    request: Request,
    project_key: str = Form(...),
    bidder_id: int = Form(...),
    document_id: int = Form(...),
    linked: str = Form("false"),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    if linked in ("true", "on", "1", "yes"):
        link_document_to_bidder(bidder_id, document_id)
    else:
        unlink_document_from_bidder(bidder_id, document_id)
    return RedirectResponse(url=f"/evaluation?project_key={project_key}", status_code=303)


@router.post("/evaluation/bidder-doc-upload", response_class=HTMLResponse)
async def evaluation_bidder_doc_upload(
    request: Request,
    project_key: str = Form(...),
    bidder_id: int = Form(...),
    file: UploadFile = File(...),
    classification: str = Form(""),
    doc_subtype: str = Form(""),
    doc_subtypes: list[str] = Form([]),
    chunk_size: int = Form(1000),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    file_bytes = await file.read()
    redirect = f"/evaluation?project_key={project_key}"
    if not file_bytes or not file.filename:
        return RedirectResponse(url=f"{redirect}&doc_upload=error", status_code=303)

    cls = (classification or "").strip() or ANGEbot_CLASSIFICATION
    subtype = (doc_subtype or "").strip() or None
    if subtype and subtype not in ANGEbot_SUBTYPES:
        subtype = None
    subtypes = [
        s.strip() for s in doc_subtypes
        if s.strip() in ANGEbot_SUBTYPES
    ]
    if not subtypes and subtype:
        subtypes = [subtype]
    # Erster Subtyp für Chunk-Prefix beim Ingest (Document.doc_subtype bleibt optional)
    ingest_subtype = subtypes[0] if subtypes else subtype
    chunk_size = normalize_chunk_size(
        chunk_size,
        fallback=recommended_chunk_size(
            cls, tender_role=None, filename=file.filename or "",
        ),
    )

    success, _msg = ingest_document(
        file_name=file.filename,
        file_bytes=file_bytes,
        classification=cls,
        chunk_size=chunk_size,
        doc_subtype=ingest_subtype,
    )
    doc = get_document_by_sha256(calculate_sha256(file_bytes), include_deleted=True)
    if not doc:
        return RedirectResponse(url=f"{redirect}&doc_upload=error", status_code=303)
    link_document_to_project(project_key, doc.id)
    link_document_to_bidder(bidder_id, doc.id)
    if subtypes:
        set_bidder_doc_subtypes(bidder_id, doc.id, subtypes)
    status = "ok" if success else "linked"
    return RedirectResponse(url=f"{redirect}&doc_upload={status}", status_code=303)


@router.post("/evaluation/delete-bidder", response_class=HTMLResponse)
async def evaluation_delete_bidder(request: Request, project_key: str = Form(...), bidder_id: int = Form(...)):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    soft_delete_bidder(bidder_id)
    return RedirectResponse(url=f"/evaluation?project_key={project_key}", status_code=303)


@router.post("/evaluation/config", response_class=HTMLResponse)
async def evaluation_save_config(
    request: Request,
    project_key: str = Form(...),
    price_years: str = Form(""),
    vergabe_notes: str = Form(""),
    rag_chunks_per_role: int = Form(12),
    price_formula: str = Form("reciprocal"),
    visual_llm_provider: str = Form(""),
    visual_llm_model: str = Form(""),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    years: list[int] = []
    for part in (price_years or "").replace(";", ",").split(","):
        part = part.strip()
        if part.isdigit():
            years.append(int(part))
    save_evaluation_config(
        project_key,
        price_years=years or None,
        vergabe_notes=vergabe_notes,
        rag_chunks_per_role=rag_chunks_per_role,
        price_formula=price_formula,
        vorgaben_ki_provider=visual_llm_provider,
        vorgaben_ki_model=visual_llm_model,
    )
    return RedirectResponse(url=f"/evaluation?project_key={project_key}&config_saved=1", status_code=303)


@router.post("/evaluation/rechunk-doc", response_class=HTMLResponse)
async def evaluation_rechunk_doc(
    request: Request,
    project_key: str = Form(...),
    document_id: int = Form(...),
    chunk_size: int = Form(1000),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    chunk_size = normalize_chunk_size(chunk_size, fallback=1000)
    ok, _msg = await asyncio.to_thread(force_rechunk_document, document_id, chunk_size)
    status = "ok" if ok else "error"
    return RedirectResponse(
        url=f"/evaluation?project_key={project_key}&rechunk={status}",
        status_code=303,
    )


@router.get("/evaluation/extract-criteria")
async def evaluation_extract_criteria_get(project_key: str = ""):
    """GET auf POST-URL (Lesezeichen/Reload) → zurück zur Offert-Seite."""
    url = f"/evaluation?project_key={project_key}" if project_key else "/evaluation"
    return RedirectResponse(url=url, status_code=303)


@router.get("/evaluation/criteria-preview", response_class=HTMLResponse)
async def evaluation_criteria_preview_page(
    request: Request,
    project_key: str,
    preview_id: str,
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    body = load_criteria_preview(preview_id, project_key)
    if not body:
        return RedirectResponse(
            url=f"/evaluation?project_key={project_key}&criteria_preview=expired",
            status_code=303,
        )
    display = {k: v for k, v in body.items() if k not in ("project_key", "stored_at")}
    return templates.TemplateResponse(
        "evaluation/_criteria_extract.html",
        _criteria_preview_template_ctx(request, project_key, display),
    )


@router.post("/evaluation/extract-criteria", response_class=HTMLResponse)
async def evaluation_extract_criteria(
    request: Request,
    project_key: str = Form(...),
    provider: str = Form(""),
    model: str = Form(""),
    visual_llm_provider: str = Form(""),
    visual_llm_model: str = Form(""),
    cloud_confirm: str = Form("false"),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    settings = load_user_settings()
    gp = (provider or settings.get("provider") or "openai").strip()
    gm = (model or settings.get("model") or "").strip()
    ap, am = resolve_vorgaben_ki(
        project_key, visual_llm_provider, visual_llm_model,
        global_provider=gp, global_model=gm,
    )
    gate = validate_tender_cloud_gate(ap, project_key, cloud_confirm in ("true", "on", "1", "yes"))
    if gate:
        return _criteria_preview_redirect(
            project_key,
            error="cloud_confirm",
            payload={},
            criteria_json="{}",
            warnings=[],
            preview_meta={},
            apply_error=None,
        )
    result = await asyncio.to_thread(
        extract_criteria_from_tender_docs, project_key, ap, am or None
    )
    payload = result.get("payload") or {}
    warnings = result.get("warnings") or []
    if not warnings and payload:
        from src.m15_evaluation import validate_criteria_payload
        warnings = validate_criteria_payload(payload)
    return _criteria_preview_redirect(
        project_key,
        error=result.get("error"),
        payload=payload,
        criteria_json=json.dumps(payload, ensure_ascii=False, indent=2),
        warnings=warnings,
        preview_meta=criteria_preview_meta(payload) if payload else {},
        apply_error=None,
        raw_llm=result.get("raw_llm"),
    )


@router.post("/evaluation/apply-criteria", response_class=HTMLResponse)
async def evaluation_apply_criteria(
    request: Request,
    project_key: str = Form(...),
    criteria_json: str = Form(...),
    confirm_apply: str = Form("false"),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    try:
        data = json.loads(criteria_json or "{}")
    except json.JSONDecodeError:
        return RedirectResponse(
            url=f"/evaluation?project_key={project_key}&criteria_apply=invalid",
            status_code=303,
        )
    from src.m15_evaluation import validate_criteria_payload

    warnings = validate_criteria_payload(data)
    confirmed = confirm_apply in ("true", "on", "1", "yes")
    if criteria_apply_requires_confirm(data) and not confirmed:
        return _criteria_preview_redirect(
            project_key,
            error=None,
            apply_error="confirm_required",
            payload=data,
            criteria_json=json.dumps(data, ensure_ascii=False, indent=2),
            warnings=warnings,
            preview_meta=criteria_preview_meta(data),
        )
    stats = import_criteria_payload(project_key, data, skip_existing=True)
    warn_q = ""
    if stats.get("warnings"):
        from urllib.parse import quote
        warn_q = "&warn=" + quote("; ".join(stats["warnings"][:5]))
    return RedirectResponse(
        url=(
            f"/evaluation?project_key={project_key}"
            f"&criteria_apply=ok&created={stats['created']}&skipped={stats['skipped']}{warn_q}"
        ),
        status_code=303,
    )


@router.post("/evaluation/delete-criterion", response_class=HTMLResponse)
async def evaluation_delete_criterion(
    request: Request,
    project_key: str = Form(...),
    criterion_id: int = Form(...),
    confirm_active_evaluation: str = Form(""),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    confirmed = confirm_active_evaluation.lower() in ("true", "on", "1", "yes")
    try:
        validate_criterion_change_during_evaluation(
            project_key, confirm_active_evaluation=confirmed, structural=True,
        )
    except ValueError:
        return RedirectResponse(
            url=f"/evaluation?project_key={project_key}&criteria_change=confirm_required",
            status_code=303,
        )
    soft_delete_criterion(criterion_id)
    return RedirectResponse(url=f"/evaluation?project_key={project_key}", status_code=303)


# ── Preisblatt (TCO) ────────────────────────────────────────────────────────

@router.get("/evaluation/price", response_class=HTMLResponse)
async def evaluation_price_page(request: Request, project_key: str, bidder_id: int = 0):
    who = _username(request)
    bidders = list_bidders(project_key)
    if not bidder_id and bidders:
        bidder_id = bidders[0].id
    items = list_price_items(bidder_id) if bidder_id else []
    einmalig = [i for i in items if i.category == "einmalig"]
    wiederkehrend = [i for i in items if i.category == "wiederkehrend"]
    by_year: dict[int, list] = {}
    for i in wiederkehrend:
        by_year.setdefault(i.year, []).append(i)
    tco = compute_bidder_tco(bidder_id) if bidder_id else None
    has_price_template = bool(
        get_tender_document_ids(project_key, roles=("preisblatt_vorlage",))
    ) if project_key else False
    has_bidder_preisblatt = bool(get_bidder_preisblatt_doc_ids(bidder_id)) if bidder_id else False
    eval_config = get_evaluation_config(project_key) if project_key else {}
    price_status = price_offers_status(project_key) if project_key else {}
    price_page_ctx = {
            "request": request,
            "project_key": project_key,
            "project_title": _project_title(project_key),
            "bidders": bidders,
            "bidder_id": bidder_id,
            "einmalig": einmalig,
            "by_year": dict(sorted(by_year.items())),
            "tco": tco,
            "may_evaluate": can_evaluate(who),
            "has_price_template": has_price_template,
            "has_bidder_preisblatt": has_bidder_preisblatt,
            "price_message": request.query_params.get("price_msg"),
            "eval_config": eval_config,
            "price_years": eval_config.get("price_years", []),
            "price_offers_status": price_status,
    }
    price_page_ctx.update(_llm_picker_context())
    return templates.TemplateResponse("evaluation/_price.html", price_page_ctx)


@router.post("/evaluation/price-item", response_class=HTMLResponse)
async def evaluation_save_price_item(
    request: Request,
    project_key: str = Form(...),
    bidder_id: int = Form(...),
    item_id: str = Form(""),
    category: str = Form(...),
    leistungsbeschreibung: str = Form(...),
    anzahl: float = Form(0.0),
    kosten_pro_einheit: float = Form(0.0),
    year: str = Form(""),
    einheit: str = Form(""),
    referenz: str = Form(""),
    bemerkung: str = Form(""),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    upsert_price_item(
        int(item_id) if item_id.strip() else None,
        bidder_id,
        category,
        leistungsbeschreibung,
        anzahl=anzahl,
        kosten_pro_einheit=kosten_pro_einheit,
        year=int(year) if year.strip() else None,
        einheit=einheit or None,
        referenz=referenz or None,
        bemerkung=bemerkung or None,
    )
    sync_result = sync_price_criterion_scores(project_key)
    extra = ""
    if not sync_result.get("synced"):
        extra = "&price_msg=offers_incomplete"
    return RedirectResponse(
        url=f"/evaluation/price?project_key={project_key}&bidder_id={bidder_id}{extra}",
        status_code=303,
    )


@router.post("/evaluation/price-item/delete", response_class=HTMLResponse)
async def evaluation_delete_price_item(
    request: Request, project_key: str = Form(...), bidder_id: int = Form(...), item_id: int = Form(...)
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    delete_price_item(item_id)
    sync_result = sync_price_criterion_scores(project_key)
    extra = ""
    if not sync_result.get("synced"):
        extra = "&price_msg=offers_incomplete"
    return RedirectResponse(
        url=f"/evaluation/price?project_key={project_key}&bidder_id={bidder_id}{extra}",
        status_code=303,
    )


@router.post("/evaluation/price-template/seed", response_class=HTMLResponse)
async def evaluation_price_template_seed(
    request: Request,
    project_key: str = Form(...),
    bidder_id: int = Form(...),
    provider: str = Form(""),
    model: str = Form(""),
    visual_llm_provider: str = Form(""),
    visual_llm_model: str = Form(""),
    cloud_confirm: str = Form("false"),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    settings = load_user_settings()
    gp = (provider or settings.get("provider") or "openai").strip()
    gm = (model or settings.get("model") or "").strip()
    ap, am = resolve_vorgaben_ki(
        project_key, visual_llm_provider, visual_llm_model,
        global_provider=gp, global_model=gm,
    )
    gate = validate_tender_cloud_gate(ap, project_key, cloud_confirm in ("true", "on", "1", "yes"))
    if gate:
        return RedirectResponse(
            url=f"/evaluation/price?project_key={project_key}&bidder_id={bidder_id}&price_msg=cloud_confirm",
            status_code=303,
        )
    extracted = await asyncio.to_thread(
        extract_price_structure_from_tender, project_key, ap, am or None
    )
    if extracted.get("error"):
        msg = "template_error"
    else:
        stats = seed_price_structure_for_bidder(
            bidder_id, extracted.get("structure") or {}, only_if_empty=True
        )
        sync_result = sync_price_criterion_scores(project_key)
        msg = f"seed_ok_{stats['created']}"
        if not sync_result.get("synced"):
            msg = "offers_incomplete"
    return RedirectResponse(
        url=f"/evaluation/price?project_key={project_key}&bidder_id={bidder_id}&price_msg={msg}",
        status_code=303,
    )


@router.post("/evaluation/price-template/parse", response_class=HTMLResponse)
async def evaluation_price_template_parse(
    request: Request,
    project_key: str = Form(...),
    bidder_id: int = Form(...),
    provider: str = Form(""),
    model: str = Form(""),
    visual_llm_provider: str = Form(""),
    visual_llm_model: str = Form(""),
    cloud_confirm: str = Form("false"),
):
    if not can_evaluate(_username(request)):
        raise HTTPException(403, "Keine Berechtigung")
    settings = load_user_settings()
    gp = (provider or settings.get("provider") or "openai").strip()
    gm = (model or settings.get("model") or "").strip()
    ap, am = resolve_vorgaben_ki(
        project_key, visual_llm_provider, visual_llm_model,
        global_provider=gp, global_model=gm,
    )
    gate = validate_evaluation_cloud_gate(
        ap, bidder_id, cloud_confirm in ("true", "on", "1", "yes")
    )
    if gate:
        return RedirectResponse(
            url=f"/evaluation/price?project_key={project_key}&bidder_id={bidder_id}&price_msg=cloud_confirm",
            status_code=303,
        )
    extracted = await asyncio.to_thread(
        extract_price_from_bidder_doc, bidder_id, project_key, ap, am or None
    )
    if extracted.get("error"):
        msg = "parse_error"
    else:
        stats = merge_price_structure_for_bidder(bidder_id, extracted.get("structure") or {})
        sync_result = sync_price_criterion_scores(project_key)
        msg = f"parse_ok_{stats['created']}_{stats['updated']}"
        if not sync_result.get("synced"):
            msg = "offers_incomplete"
    return RedirectResponse(
        url=f"/evaluation/price?project_key={project_key}&bidder_id={bidder_id}&price_msg={msg}",
        status_code=303,
    )


def _export_rows(project_key: str, may_see: bool) -> tuple[list[str], list[list]]:
    sheets = build_evaluation_export_sheets(
        project_key,
        project_title=_project_title(project_key),
        may_see_evaluators=may_see,
    )
    return sheets["Bewertungen"]


@router.get("/evaluation/export.csv")
async def evaluation_export_csv(request: Request, project_key: str):
    who = _username(request)
    if not who:
        raise HTTPException(401)
    headers, rows = _export_rows(project_key, can_view_evaluator_details(who))

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=bewertung_{project_key[:24]}.csv"},
    )


@router.get("/evaluation/export.xlsx")
async def evaluation_export_xlsx(request: Request, project_key: str):
    who = _username(request)
    if not who:
        raise HTTPException(401)
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl nicht installiert")

    headers, rows = _export_rows(project_key, can_view_evaluator_details(who))
    sheets = build_evaluation_export_sheets(
        project_key,
        project_title=_project_title(project_key),
        may_see_evaluators=can_view_evaluator_details(who),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bewertungen"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    detail_headers, detail_rows = sheets.get("Einzelanforderungen", ([], []))
    if detail_rows:
        ws_detail = wb.create_sheet("Einzelanforderungen")
        ws_detail.append(detail_headers)
        for row in detail_rows:
            ws_detail.append(row)

    ws2 = wb.create_sheet("Rangfolge")
    ws2.append(["Rang", "Bieter", "Gesamt %", "KO"])
    rankings = compute_rankings(project_key)
    for r in rankings:
        ws2.append([r.get("rank"), r.get("bidder_name"), r.get("total_score"), r.get("ko")])

    if any(r.get("has_phase2") for r in rankings):
        ws3 = wb.create_sheet("Rangfolge Phase 1")
        ws3.append([
            "Rang Phase 1", "Bieter", "ZK %", "Max. bei Präsentation voll", "Einladung?", "KO",
        ])
        for r in rankings:
            invite = ""
            if r.get("can_still_win") is True:
                invite = "ja"
            elif r.get("can_still_win") is False:
                invite = "nein"
            ws3.append([
                r.get("interim_rank"),
                r.get("bidder_name"),
                r.get("interim_score"),
                r.get("max_score"),
                invite,
                r.get("ko"),
            ])

    ws_price = wb.create_sheet("Preisblatt")
    eval_cfg = get_evaluation_config(project_key)
    year_cols = eval_cfg.get("price_years") or [2027, 2028, 2029, 2030]
    ws_price.append(
        ["Bieter", "Einmalig CHF"] + [str(y) for y in year_cols]
        + ["Total exkl. MwSt", "MwSt", "Total inkl. MwSt"]
    )
    for bidder in list_bidders(project_key):
        tco = compute_bidder_tco(bidder.id)
        by_year = tco["by_year"]
        ws_price.append(
            [bidder.name, tco["einmalig_total"]]
            + [by_year.get(y, 0) for y in year_cols]
            + [tco["total_exkl_mwst"], tco["mwst"], tco["total_inkl_mwst"]]
        )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=bewertung_{project_key[:24]}.xlsx"},
    )
